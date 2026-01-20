#!/usr/bin/env python3
"""Test and fill gender sections in template."""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np
import sys

# Grade to points mapping
grade_to_points = {
    'A': 12, 'A-': 11, 'B+': 10, 'B': 9, 'B-': 8,
    'C+': 7, 'C': 6, 'C-': 5, 'D+': 4, 'D': 3, 'D-': 2, 'E': 1
}

def calculate_grade_distribution(mean_grades):
    """Calculate grade distribution from mean grades."""
    grades = mean_grades.dropna().astype(str)
    distribution = {}
    grade_order = ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']
    for grade in grade_order:
        distribution[grade] = len(grades[grades == grade])
    ab_count = distribution.get('A', 0) + distribution.get('A-', 0) + \
               distribution.get('B+', 0) + distribution.get('B', 0) + distribution.get('B-', 0)
    return distribution, ab_count

def calculate_mean_points(mean_grades):
    """Calculate mean points from grades."""
    grades = mean_grades.dropna().astype(str)
    points_list = [grade_to_points[g] for g in grades if g in grade_to_points]
    return np.mean(points_list) if points_list else None

# Load results
print("Loading results file...", file=sys.stdout, flush=True)
results_df = pd.read_excel('Students Upload KCSE Results - Template_updated.xlsx')
print(f"Total students: {len(results_df)}", file=sys.stdout, flush=True)
print(f"Columns: {list(results_df.columns)[:5]}...", file=sys.stdout, flush=True)

# Check for GENDER column
gender_col = None
for col in results_df.columns:
    if 'GENDER' in str(col).upper() or 'GEND' in str(col).upper():
        gender_col = col
        break

if not gender_col:
    print("ERROR: GENDER column not found!", file=sys.stderr, flush=True)
    print(f"Available columns: {list(results_df.columns)}", file=sys.stderr, flush=True)
    sys.exit(1)

print(f"Found gender column: '{gender_col}'", file=sys.stdout, flush=True)

# Normalize gender
def normalize_gender(val):
    if pd.isna(val):
        return None
    val_str = str(val).strip().upper()
    if val_str in ['M', 'MALE', 'BOY', 'BOYS']:
        return 'MALE'
    elif val_str in ['F', 'FEMALE', 'GIRL', 'GIRLS']:
        return 'FEMALE'
    return None

results_df['GENDER_NORM'] = results_df[gender_col].apply(normalize_gender)
print(f"Gender distribution: {results_df['GENDER_NORM'].value_counts().to_dict()}", file=sys.stdout, flush=True)

# Calculate statistics
boys_df = results_df[results_df['GENDER_NORM'] == 'MALE']
girls_df = results_df[results_df['GENDER_NORM'] == 'FEMALE']

print(f"Boys: {len(boys_df)}, Girls: {len(girls_df)}", file=sys.stdout, flush=True)

# Calculate boys stats
boys_data = None
if len(boys_df) > 0:
    boys_grades = boys_df['MEAN_GRADE'].dropna()
    boys_dist, boys_ab = calculate_grade_distribution(boys_grades)
    boys_mean = calculate_mean_points(boys_grades)
    boys_data = {'count': len(boys_df), 'grade_dist': boys_dist, 'ab_count': boys_ab, 'mean_points': boys_mean}
    print(f"Boys data calculated: {boys_data['count']} students, AB: {boys_data['ab_count']}", file=sys.stdout, flush=True)

# Calculate girls stats
girls_data = None
if len(girls_df) > 0:
    girls_grades = girls_df['MEAN_GRADE'].dropna()
    girls_dist, girls_ab = calculate_grade_distribution(girls_grades)
    girls_mean = calculate_mean_points(girls_grades)
    girls_data = {'count': len(girls_df), 'grade_dist': girls_dist, 'ab_count': girls_ab, 'mean_points': girls_mean}
    print(f"Girls data calculated: {girls_data['count']} students, AB: {girls_data['ab_count']}", file=sys.stdout, flush=True)

# Load template
print("\nLoading template...", file=sys.stdout, flush=True)
wb = load_workbook('KCSE ANALYSIS TEMPLATE.xlsx')
sheet = wb.active

# Grade column mapping
col_map = {
    'A': 7, 'A-': 8, 'B+': 9, 'B': 10, 'B-': 11,
    'C+': 12, 'C': 13, 'C-': 14, 'D+': 15, 'D': 16, 'D-': 17, 'E': 18
}

# Fill Row 7 - ORDER OF MERIT
print("Filling Row 7 (ORDER OF MERIT)...", file=sys.stdout, flush=True)
if boys_data:
    sheet.cell(row=7, column=3).value = boys_data['count']  # BOYS
    print(f"  Set BOYS (C7) = {boys_data['count']}", file=sys.stdout, flush=True)
if girls_data:
    sheet.cell(row=7, column=4).value = girls_data['count']  # GIRLS
    print(f"  Set GIRLS (D7) = {girls_data['count']}", file=sys.stdout, flush=True)

# Fill Row 13 - BOYS
if boys_data:
    print("Filling Row 13 (BOYS)...", file=sys.stdout, flush=True)
    sheet.cell(row=13, column=3).value = boys_data['count']  # BOYS column
    sheet.cell(row=13, column=5).value = boys_data['count']  # TOTAL column
    sheet.cell(row=13, column=6).value = boys_data['ab_count']  # AB
    for grade, col in col_map.items():
        sheet.cell(row=13, column=col).value = boys_data['grade_dist'].get(grade, 0)
    if boys_data['mean_points']:
        sheet.cell(row=13, column=23).value = round(boys_data['mean_points'], 2)
    print(f"  Filled BOYS row with {boys_data['count']} students", file=sys.stdout, flush=True)

# Fill Row 14 - GIRLS
if girls_data:
    print("Filling Row 14 (GIRLS)...", file=sys.stdout, flush=True)
    sheet.cell(row=14, column=4).value = girls_data['count']  # GIRLS column
    sheet.cell(row=14, column=5).value = girls_data['count']  # TOTAL column
    sheet.cell(row=14, column=6).value = girls_data['ab_count']  # AB
    for grade, col in col_map.items():
        sheet.cell(row=14, column=col).value = girls_data['grade_dist'].get(grade, 0)
    if girls_data['mean_points']:
        sheet.cell(row=14, column=23).value = round(girls_data['mean_points'], 2)
    print(f"  Filled GIRLS row with {girls_data['count']} students", file=sys.stdout, flush=True)

# Save
output_file = 'KCSE ANALYSIS TEMPLATE_2025_FILLED.xlsx'
print(f"\nSaving to {output_file}...", file=sys.stdout, flush=True)
wb.save(output_file)
print("✓ Template saved successfully!", file=sys.stdout, flush=True)

# Verify
print("\nVerification:", file=sys.stdout, flush=True)
row13 = [c.value for c in sheet[13]]
row14 = [c.value for c in sheet[14]]
print(f"Row 13 BOYS (C13): {row13[2]}", file=sys.stdout, flush=True)
print(f"Row 14 GIRLS (D14): {row14[3]}", file=sys.stdout, flush=True)

sys.stdout.flush()
