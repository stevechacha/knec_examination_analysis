#!/usr/bin/env python3
"""
Create a completely new and filled KCSE Analysis file.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
import numpy as np

# Grade to points mapping
grade_to_points = {
    'A': 12, 'A-': 11, 'B+': 10, 'B': 9, 'B-': 8,
    'C+': 7, 'C': 6, 'C-': 5, 'D+': 4, 'D': 3, 'D-': 2, 'E': 1
}

def calculate_grade_distribution(grades):
    """Calculate grade distribution."""
    g = grades.dropna().astype(str)
    g = g[g != 'nan']
    dist = {}
    for grade in ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']:
        dist[grade] = len(g[g == grade])
    ab = dist.get('A', 0) + dist.get('A-', 0) + dist.get('B+', 0) + dist.get('B', 0) + dist.get('B-', 0)
    return dist, ab

def calculate_mean_points(grades):
    """Calculate mean points."""
    g = grades.dropna().astype(str)
    g = g[g != 'nan']
    pts = [grade_to_points[x] for x in g if x in grade_to_points]
    return np.mean(pts) if pts else None

def find_gender_column(df):
    """Find gender column."""
    for col in df.columns:
        if 'GENDER' in str(col).upper():
            return col
    return None

def normalize_gender(val):
    """Normalize gender."""
    if pd.isna(val):
        return None
    v = str(val).strip().upper()
    return 'MALE' if v in ['M', 'MALE', 'BOY', 'BOYS'] else 'FEMALE' if v in ['F', 'FEMALE', 'GIRL', 'GIRLS'] else None

print("="*70)
print("CREATING COMPLETE KCSE ANALYSIS FILE - 2025")
print("="*70)

# Load data
print("\n1. Loading results...")
df = pd.read_excel('Students Upload KCSE Results - Template_updated.xlsx')
total = len(df)
print(f"   ✓ {total} students loaded")

# Find gender
gender_col = find_gender_column(df)
if gender_col:
    df['GEND'] = df[gender_col].apply(normalize_gender)
    boys_df = df[df['GEND'] == 'MALE'].copy()
    girls_df = df[df['GEND'] == 'FEMALE'].copy()
    print(f"   ✓ Boys: {len(boys_df)}, Girls: {len(girls_df)}")
else:
    boys_df = pd.DataFrame()
    girls_df = pd.DataFrame()
    print("   ⚠ No gender data")

# Overall stats
mg = df['MEAN_GRADE'].dropna()
gd, ab = calculate_grade_distribution(mg)
mp = calculate_mean_points(mg)
print(f"   ✓ Overall: AB={ab}, Mean={mp:.2f}" if mp else f"   ✓ Overall: AB={ab}")

# Gender stats
boys_data = None
if len(boys_df) > 0:
    bg = boys_df['MEAN_GRADE'].dropna()
    bd, bab = calculate_grade_distribution(bg)
    bmp = calculate_mean_points(bg)
    boys_data = {'count': len(boys_df), 'dist': bd, 'ab': bab, 'mean': bmp}
    print(f"   ✓ Boys: AB={bab}, Mean={bmp:.2f}" if bmp else f"   ✓ Boys: AB={bab}")

girls_data = None
if len(girls_df) > 0:
    gg = girls_df['MEAN_GRADE'].dropna()
    gd_girls, gab = calculate_grade_distribution(gg)
    gmp = calculate_mean_points(gg)
    girls_data = {'count': len(girls_df), 'dist': gd_girls, 'ab': gab, 'mean': gmp}
    print(f"   ✓ Girls: AB={gab}, Mean={gmp:.2f}" if gmp else f"   ✓ Girls: AB={gab}")

# Subject mapping
subject_map = {
    'ENG': ('English', '101'),
    'KIS': ('Kiswahili', '102'),
    'MAT': ('Mathematics', '121'),
    'BIO': ('Biology', '231'),
    'PHY': ('Physics', '232'),
    'CHE': ('Chemistry', '233'),
    'HIS': ('History and Government', '311'),
    'GEO': ('Geography', '312'),
    'CRE': ('C.R.E.', '313'),
    'AGR': ('Agriculture', '443'),
    'BST': ('Business Studies', '565'),
    'COM': ('Computer Studies', '451')
}

# Calculate subject stats
print("\n2. Calculating subject statistics...")
subject_stats = {}
for subj, (name, code) in subject_map.items():
    if subj in df.columns:
        sg = df[subj].dropna().astype(str)
        sg = sg[sg != 'nan']
        if len(sg) > 0:
            sd, sab = calculate_grade_distribution(sg)
            smp = calculate_mean_points(sg)
            subject_stats[subj] = {
                'name': name, 'code': code,
                'total': len(sg), 'dist': sd, 'ab': sab, 'mean': smp
            }

# Subject stats by gender
boys_subj_stats = {}
if len(boys_df) > 0:
    for subj, (name, code) in subject_map.items():
        if subj in boys_df.columns:
            sg = boys_df[subj].dropna().astype(str)
            sg = sg[sg != 'nan']
            if len(sg) > 0:
                sd, sab = calculate_grade_distribution(sg)
                smp = calculate_mean_points(sg)
                boys_subj_stats[subj] = {
                    'name': name, 'code': code,
                    'count': len(sg), 'dist': sd, 'ab': sab, 'mean': smp
                }

girls_subj_stats = {}
if len(girls_df) > 0:
    for subj, (name, code) in subject_map.items():
        if subj in girls_df.columns:
            sg = girls_df[subj].dropna().astype(str)
            sg = sg[sg != 'nan']
            if len(sg) > 0:
                sd, sab = calculate_grade_distribution(sg)
                smp = calculate_mean_points(sg)
                girls_subj_stats[subj] = {
                    'name': name, 'code': code,
                    'count': len(sg), 'dist': sd, 'ab': sab, 'mean': smp
                }

print(f"   ✓ Calculated stats for {len(subject_stats)} subjects")

# Create new workbook
print("\n3. Creating new Excel workbook...")
wb = Workbook()
sheet = wb.active
sheet.title = "KCSE Analysis 2025"

# Column mapping
col_map = {
    'A': 8, 'A-': 9, 'B+': 10, 'B': 11, 'B-': 12,
    'C+': 13, 'C': 14, 'C-': 15, 'D+': 16, 'D': 17, 'D-': 18, 'E': 19
}
AB_COL = 7  # Column G
YEAR_2025_COL = 25  # Column Y

# Fill ORDER OF MERIT section
print("\n4. Filling ORDER OF MERIT section...")
sheet.cell(3, 1).value = "ORDER OF MERIT"
sheet.cell(5, 1).value = "CANDIDATURE"
sheet.cell(5, 7).value = "MEAN GRADE DISTRIBUTION"
sheet.cell(5, 25).value = "MEAN SCORES"

# Headers row 6
sheet.cell(6, 1).value = "SUB COUNTY"
sheet.cell(6, 2).value = "SCHOOL"
sheet.cell(6, 3).value = "BOYS"
sheet.cell(6, 4).value = "GIRLS"
sheet.cell(6, 5).value = "TOTAL"
sheet.cell(6, 7).value = "AB"
sheet.cell(6, 8).value = "A"
sheet.cell(6, 9).value = "A-"
sheet.cell(6, 10).value = "B+"
sheet.cell(6, 11).value = "B"
sheet.cell(6, 12).value = "B-"
sheet.cell(6, 13).value = "C+"
sheet.cell(6, 14).value = "C"
sheet.cell(6, 15).value = "C-"
sheet.cell(6, 16).value = "D+"
sheet.cell(6, 17).value = "D"
sheet.cell(6, 18).value = "D-"
sheet.cell(6, 19).value = "E"
sheet.cell(6, 25).value = "2025"

# Row 7 data
sheet.cell(7, 1).value = "MABERA"
sheet.cell(7, 2).value = "KUBWEYE SECONDARY SCHOOL"
if boys_data:
    sheet.cell(7, 3).value = boys_data['count']
if girls_data:
    sheet.cell(7, 4).value = girls_data['count']
sheet.cell(7, 5).value = total
sheet.cell(7, AB_COL).value = ab
for grade, col in col_map.items():
    sheet.cell(7, col).value = gd.get(grade, 0)
if mp:
    sheet.cell(7, YEAR_2025_COL).value = round(mp, 2)
print("   ✓ Row 7 filled")

# Fill PERFORMANCE BY GENDER section
print("\n5. Filling PERFORMANCE BY GENDER section...")
sheet.cell(9, 1).value = "PERFORMANCE BY GENDER"
sheet.cell(11, 1).value = "CANDIDATURE"
sheet.cell(11, 7).value = "MEAN GRADE DISTRIBUTION"
sheet.cell(11, 25).value = "MEAN SCORES"

sheet.cell(12, 1).value = "SCHOOL"
sheet.cell(12, 2).value = "GENDER"
sheet.cell(12, 3).value = "BOYS"
sheet.cell(12, 4).value = "GIRLS"
sheet.cell(12, 5).value = "TOTAL"
sheet.cell(12, 7).value = "AB"
for i, grade in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    sheet.cell(12, i).value = grade
sheet.cell(12, 25).value = "2025"

# Row 13 - BOYS
if boys_data:
    sheet.cell(13, 1).value = "KUBWEYE SECONDARY SCHOOL"
    sheet.cell(13, 2).value = "BOYS"
    sheet.cell(13, 3).value = boys_data['count']
    sheet.cell(13, 5).value = boys_data['count']
    sheet.cell(13, AB_COL).value = boys_data['ab']
    for grade, col in col_map.items():
        sheet.cell(13, col).value = boys_data['dist'].get(grade, 0)
    if boys_data['mean']:
        sheet.cell(13, YEAR_2025_COL).value = round(boys_data['mean'], 2)
    print("   ✓ Row 13 (BOYS) filled")

# Row 14 - GIRLS
if girls_data:
    sheet.cell(14, 1).value = "KUBWEYE SECONDARY SCHOOL"
    sheet.cell(14, 2).value = "GIRLS"
    sheet.cell(14, 4).value = girls_data['count']
    sheet.cell(14, 5).value = girls_data['count']
    sheet.cell(14, AB_COL).value = girls_data['ab']
    for grade, col in col_map.items():
        sheet.cell(14, col).value = girls_data['dist'].get(grade, 0)
    if girls_data['mean']:
        sheet.cell(14, YEAR_2025_COL).value = round(girls_data['mean'], 2)
    print("   ✓ Row 14 (GIRLS) filled")

# Row 15 - TOTAL
sheet.cell(15, 2).value = "TOTAL"
sheet.cell(15, 5).value = total
sheet.cell(15, AB_COL).value = ab
for grade, col in col_map.items():
    sheet.cell(15, col).value = gd.get(grade, 0)
if mp:
    sheet.cell(15, YEAR_2025_COL).value = round(mp, 2)
print("   ✓ Row 15 (TOTAL) filled")

# Fill GENERAL SUBJECT ANALYSIS
print("\n6. Filling GENERAL SUBJECT ANALYSIS...")
sheet.cell(17, 1).value = "SUBJECT ANALYSIS BY IMPROVEMENT INDEX"
sheet.cell(19, 1).value = "CANDIDATURE"
sheet.cell(19, 7).value = "MEAN GRADE DISTRIBUTION"
sheet.cell(19, 25).value = "MEAN SCORES"

sheet.cell(20, 1).value = "SUBJECT"
sheet.cell(20, 2).value = "CODE"
sheet.cell(20, 3).value = "BOYS"
sheet.cell(20, 4).value = "GIRLS"
sheet.cell(20, 5).value = "TOTAL"
sheet.cell(20, 7).value = "AB"
for i, grade in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    sheet.cell(20, i).value = grade
sheet.cell(20, 25).value = "2025"

row = 40
for subj, (name, code) in subject_map.items():
    if subj in subject_stats:
        stats = subject_stats[subj]
        sheet.cell(row, 2).value = stats['name']
        sheet.cell(row, 3).value = stats['code']
        sheet.cell(row, 6).value = stats['total']
        sheet.cell(row, AB_COL).value = stats['ab']
        for grade, col in col_map.items():
            sheet.cell(row, col).value = stats['dist'].get(grade, 0)
        if stats['mean']:
            sheet.cell(row, YEAR_2025_COL).value = round(stats['mean'], 2)
        row += 1
print(f"   ✓ Filled {len(subject_stats)} subjects")

# Fill SUBJECT ANALYSIS (BOYS)
print("\n7. Filling SUBJECT ANALYSIS (BOYS)...")
sheet.cell(36, 1).value = "GENERAL SUBJECT ANALYSIS"
sheet.cell(38, 1).value = "CANDIDATURE"
sheet.cell(38, 7).value = "MEAN GRADE DISTRIBUTION"
sheet.cell(38, 25).value = "MEAN SCORES"

sheet.cell(39, 1).value = "SUBJECT"
sheet.cell(39, 2).value = "CODE"
sheet.cell(39, 3).value = "BOYS"
sheet.cell(39, 4).value = "GIRLS"
sheet.cell(39, 5).value = "TOTAL"
sheet.cell(39, 7).value = "AB"
for i, grade in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    sheet.cell(39, i).value = grade
sheet.cell(39, 25).value = "2025"

row = 56
for subj, (name, code) in subject_map.items():
    if subj in boys_subj_stats:
        stats = boys_subj_stats[subj]
        sheet.cell(row, 2).value = stats['name']
        sheet.cell(row, 3).value = stats['code']
        sheet.cell(row, 4).value = stats['count']
        sheet.cell(row, 6).value = stats['count']
        sheet.cell(row, AB_COL).value = stats['ab']
        for grade, col in col_map.items():
            sheet.cell(row, col).value = stats['dist'].get(grade, 0)
        if stats['mean']:
            sheet.cell(row, YEAR_2025_COL).value = round(stats['mean'], 2)
        row += 1
print(f"   ✓ Filled {len(boys_subj_stats)} subjects for boys")

# Fill SUBJECT ANALYSIS (GIRLS)
print("\n8. Filling SUBJECT ANALYSIS (GIRLS)...")
sheet.cell(52, 1).value = "SUBJECT ANALYSIS (BOYS)"
sheet.cell(54, 1).value = "CANDIDATURE"
sheet.cell(54, 7).value = "MEAN GRADE DISTRIBUTION"
sheet.cell(54, 25).value = "MEAN SCORES"

sheet.cell(55, 1).value = "SUBJECT"
sheet.cell(55, 2).value = "CODE"
sheet.cell(55, 3).value = "BOYS"
sheet.cell(55, 4).value = "GIRLS"
sheet.cell(55, 5).value = "TOTAL"
sheet.cell(55, 7).value = "AB"
for i, grade in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    sheet.cell(55, i).value = grade
sheet.cell(55, 25).value = "2025"

sheet.cell(70, 1).value = "SUBJECT ANALYSIS (GIRLS)"
sheet.cell(72, 1).value = "CANDIDATURE"
sheet.cell(72, 7).value = "MEAN GRADE DISTRIBUTION"
sheet.cell(72, 25).value = "MEAN SCORES"

sheet.cell(73, 1).value = "SUBJECT"
sheet.cell(73, 2).value = "CODE"
sheet.cell(73, 3).value = "BOYS"
sheet.cell(73, 4).value = "GIRLS"
sheet.cell(73, 5).value = "TOTAL"
sheet.cell(73, 7).value = "AB"
for i, grade in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    sheet.cell(73, i).value = grade
sheet.cell(73, 25).value = "2025"

row = 74
for subj, (name, code) in subject_map.items():
    if subj in girls_subj_stats:
        stats = girls_subj_stats[subj]
        sheet.cell(row, 2).value = stats['name']
        sheet.cell(row, 3).value = stats['code']
        sheet.cell(row, 5).value = stats['count']
        sheet.cell(row, 6).value = stats['count']
        sheet.cell(row, AB_COL).value = stats['ab']
        for grade, col in col_map.items():
            sheet.cell(row, col).value = stats['dist'].get(grade, 0)
        if stats['mean']:
            sheet.cell(row, YEAR_2025_COL).value = round(stats['mean'], 2)
        row += 1
print(f"   ✓ Filled {len(girls_subj_stats)} subjects for girls")

# Fill TOP 10 BOYS
print("\n9. Filling TOP 10 BOYS...")
sheet.cell(88, 1).value = "TOP 10 BOYS"
sheet.cell(89, 1).value = "INDEX_NO"
sheet.cell(89, 2).value = "NAME"
for i, subj in enumerate(['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'HIS', 'GEO', 'CRE', 'AGR', 'BST'], start=3):
    sheet.cell(89, i).value = subj
sheet.cell(89, 18).value = "GR"

if len(boys_df) > 0:
    boys_df['MEAN_PTS'] = boys_df['MEAN_GRADE'].apply(lambda x: grade_to_points.get(str(x).strip(), 0) if pd.notna(x) else 0)
    top_boys = boys_df.nlargest(10, 'MEAN_PTS')
    
    row = 90
    for idx, student in top_boys.iterrows():
        sheet.cell(row, 1).value = str(student['INDEXNO'])
        sheet.cell(row, 2).value = student['NAME']
        for i, subj in enumerate(['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'HIS', 'GEO', 'CRE', 'AGR', 'BST'], start=3):
            if subj in student and pd.notna(student[subj]) and str(student[subj]) != 'nan':
                sheet.cell(row, i).value = str(student[subj])
        sheet.cell(row, 18).value = student['MEAN_GRADE']
        row += 1
    print(f"   ✓ Filled top {len(top_boys)} boys")

# Fill TOP 10 GIRLS
print("\n10. Filling TOP 10 GIRLS...")
sheet.cell(103, 1).value = "TOP 10 GIRLS"
sheet.cell(104, 1).value = "INDEX_NO"
sheet.cell(104, 2).value = "NAME"
for i, subj in enumerate(['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'HIS', 'GEO', 'CRE', 'AGR', 'BST'], start=3):
    sheet.cell(104, i).value = subj
sheet.cell(104, 18).value = "GR"

if len(girls_df) > 0:
    girls_df['MEAN_PTS'] = girls_df['MEAN_GRADE'].apply(lambda x: grade_to_points.get(str(x).strip(), 0) if pd.notna(x) else 0)
    top_girls = girls_df.nlargest(10, 'MEAN_PTS')
    
    row = 105
    for idx, student in top_girls.iterrows():
        sheet.cell(row, 1).value = str(student['INDEXNO'])
        sheet.cell(row, 2).value = student['NAME']
        for i, subj in enumerate(['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'HIS', 'GEO', 'CRE', 'AGR', 'BST'], start=3):
            if subj in student and pd.notna(student[subj]) and str(student[subj]) != 'nan':
                sheet.cell(row, i).value = str(student[subj])
        sheet.cell(row, 18).value = student['MEAN_GRADE']
        row += 1
    print(f"   ✓ Filled top {len(top_girls)} girls")

# Save with new name
print("\n11. Saving file...")
output_file = 'KCSE_ANALYSIS_2025_COMPLETE.xlsx'
wb.save(output_file)
print(f"   ✓ Saved to: {output_file}")

print("\n" + "="*70)
print("SUCCESS! Complete KCSE Analysis file created!")
print("="*70)
print(f"\nFile: {output_file}")
print(f"Total Students: {total}")
if boys_data:
    print(f"Boys: {boys_data['count']} (AB: {boys_data['ab']}, Mean: {boys_data['mean']:.2f})")
if girls_data:
    print(f"Girls: {girls_data['count']} (AB: {girls_data['ab']}, Mean: {girls_data['mean']:.2f})")
print(f"Subjects analyzed: {len(subject_stats)}")
print(f"Overall Mean Points: {mp:.2f}" if mp else "Overall Mean Points: N/A")
print("="*70)
