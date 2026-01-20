#!/usr/bin/env python3
"""
Script to fill KCSE Analysis Template with 2025 results data.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

# Grade to points mapping (KCSE system)
grade_to_points = {
    'A': 12, 'A-': 11,
    'B+': 10, 'B': 9, 'B-': 8,
    'C+': 7, 'C': 6, 'C-': 5,
    'D+': 4, 'D': 3, 'D-': 2,
    'E': 1
}

def calculate_grade_distribution(mean_grades):
    """Calculate grade distribution from mean grades."""
    grades = mean_grades.dropna().astype(str)
    distribution = {}
    
    grade_order = ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']
    for grade in grade_order:
        distribution[grade] = len(grades[grades == grade])
    
    # Calculate AB (A and B bands combined)
    ab_count = distribution.get('A', 0) + distribution.get('A-', 0) + \
               distribution.get('B+', 0) + distribution.get('B', 0) + distribution.get('B-', 0)
    
    return distribution, ab_count

def calculate_mean_points(mean_grades):
    """Calculate mean points from grades."""
    grades = mean_grades.dropna().astype(str)
    points_list = []
    
    for grade in grades:
        if grade in grade_to_points:
            points_list.append(grade_to_points[grade])
    
    if points_list:
        return np.mean(points_list)
    return None

def find_gender_column(df):
    """Find the gender column in the dataframe."""
    possible_names = ['GENDER', 'Gender', 'gender', 'SEX', 'Sex', 'sex', 
                      'GEND', 'Gend', 'gend', 'M/F', 'M_F', 'MALE/FEMALE']
    
    for col in df.columns:
        col_str = str(col).strip()
        if col_str in possible_names:
            return col
        col_lower = col_str.lower()
        if 'gender' in col_lower or 'sex' in col_lower or 'gend' in col_lower:
            return col
    return None

def normalize_gender(gender_value):
    """Normalize gender values to MALE/FEMALE."""
    if pd.isna(gender_value):
        return None
    
    gender_str = str(gender_value).strip().upper()
    
    # Handle single letter M/F
    if gender_str in ['M', 'MALE', 'BOY', 'BOYS', 'MALE']:
        return 'MALE'
    elif gender_str in ['F', 'FEMALE', 'GIRL', 'GIRLS', 'FEMALE']:
        return 'FEMALE'
    return None

def fill_analysis_template():
    """Fill the KCSE Analysis Template with data."""
    
    # Load the results data
    results_df = pd.read_excel('Students Upload KCSE Results - Template_updated.xlsx')
    
    # Load the template
    wb = load_workbook('KCSE ANALYSIS TEMPLATE.xlsx')
    sheet = wb.active
    
    # Find gender column
    gender_col = find_gender_column(results_df)
    if gender_col:
        print(f"Found gender column: '{gender_col}'")
        results_df['GENDER_NORM'] = results_df[gender_col].apply(normalize_gender)
        has_gender_data = True
        print(f"Gender distribution: {results_df['GENDER_NORM'].value_counts().to_dict()}")
    else:
        has_gender_data = False
        print("Warning: Gender column not found. Gender-specific sections will not be filled.")
        print(f"Available columns: {list(results_df.columns)}")
    
    # Calculate statistics
    total_students = len(results_df)
    mean_grades = results_df['MEAN_GRADE'].dropna()
    
    # Calculate grade distribution for all students
    grade_dist, ab_count = calculate_grade_distribution(mean_grades)
    
    # Calculate mean points
    mean_points = calculate_mean_points(mean_grades)
    
    # Calculate gender-specific statistics
    boys_data = None
    girls_data = None
    
    if has_gender_data:
        boys_df = results_df[results_df['GENDER_NORM'] == 'MALE']
        girls_df = results_df[results_df['GENDER_NORM'] == 'FEMALE']
        
        if len(boys_df) > 0:
            boys_mean_grades = boys_df['MEAN_GRADE'].dropna()
            boys_grade_dist, boys_ab_count = calculate_grade_distribution(boys_mean_grades)
            boys_mean_points = calculate_mean_points(boys_mean_grades)
            boys_data = {
                'count': len(boys_df),
                'grade_dist': boys_grade_dist,
                'ab_count': boys_ab_count,
                'mean_points': boys_mean_points
            }
        
        if len(girls_df) > 0:
            girls_mean_grades = girls_df['MEAN_GRADE'].dropna()
            girls_grade_dist, girls_ab_count = calculate_grade_distribution(girls_mean_grades)
            girls_mean_points = calculate_mean_points(girls_mean_grades)
            girls_data = {
                'count': len(girls_df),
                'grade_dist': girls_grade_dist,
                'ab_count': girls_ab_count,
                'mean_points': girls_mean_points
            }
    
    # Grade distribution mapping to columns (row 6 headers: A, A-, B+, B, B-, C+, C, C-, D+, D, D-, E)
    # Column mapping: Col A=1 (SUB COUNTY), B=2 (SCHOOL), C=3 (BOYS), D=4 (GIRLS), 
    # E=5 (TOTAL), F=6 (AB), G=7 (A), H=8 (A-), I=9 (B+), J=10 (B), K=11 (B-),
    # L=12 (C+), M=13 (C), N=14 (C-), O=15 (D+), P=16 (D), Q=17 (D-), R=18 (E),
    # S=19 (X), T=20 (Y), U=21 (2023), V=22 (2024), W=23 (2025), X=24 (DEV)
    
    # Fill ORDER OF MERIT section (Row 7)
    row = 7
    # SUB COUNTY - leave as is or update if known
    # sheet.cell(row=row, column=1).value = "MABERA"  # Uncomment if needed
    
    # SCHOOL - will be filled based on screenshot data
    sheet.cell(row=row, column=2).value = "KUBWEYE SECONDARY SCHOOL"
    
    # BOYS, GIRLS - fill if gender data available
    if boys_data:
        sheet.cell(row=row, column=3).value = boys_data['count']  # BOYS
    if girls_data:
        sheet.cell(row=row, column=4).value = girls_data['count']  # GIRLS
    
    # TOTAL
    sheet.cell(row=row, column=5).value = total_students
    
    # AB (Aggregate A and B bands)
    sheet.cell(row=row, column=6).value = ab_count
    
    # Grade distribution (A, A-, B+, B, B-, C+, C, C-, D+, D, D-, E)
    col_map = {
        'A': 7, 'A-': 8, 'B+': 9, 'B': 10, 'B-': 11,
        'C+': 12, 'C': 13, 'C-': 14, 'D+': 15, 'D': 16, 'D-': 17, 'E': 18
    }
    
    for grade, col in col_map.items():
        sheet.cell(row=row, column=col).value = grade_dist.get(grade, 0)
    
    # Mean score for 2025 (column 23)
    if mean_points:
        sheet.cell(row=row, column=23).value = round(mean_points, 2)
    
    # Fill PERFORMANCE BY GENDER section (Rows 13-15)
    # Row 13: BOYS
    # Row 14: GIRLS
    # Row 15: TOTAL
    
    # Fill BOYS row (13) if data available
    if boys_data:
        row_boys = 13
        print(f"Filling BOYS row ({row_boys}): {boys_data['count']} students")
        print(f"  Setting C{row_boys} (BOYS) = {boys_data['count']}")
        print(f"  Setting E{row_boys} (TOTAL) = {boys_data['count']}")
        print(f"  Setting F{row_boys} (AB) = {boys_data['ab_count']}")
        sheet.cell(row=row_boys, column=3).value = boys_data['count']  # BOYS column (C)
        sheet.cell(row=row_boys, column=5).value = boys_data['count']  # TOTAL column (E)
        sheet.cell(row=row_boys, column=6).value = boys_data['ab_count']  # AB column (F)
        
        for grade, col in col_map.items():
            count = boys_data['grade_dist'].get(grade, 0)
            if count > 0:
                sheet.cell(row=row_boys, column=col).value = count
        
        if boys_data['mean_points']:
            sheet.cell(row=row_boys, column=23).value = round(boys_data['mean_points'], 2)
            print(f"  Setting W{row_boys} (Mean Points) = {round(boys_data['mean_points'], 2)}")
        print(f"  ✓ BOYS row filled successfully")
    else:
        print("No boys data available - BOYS row not filled")
    
    # Fill GIRLS row (14) if data available
    if girls_data:
        row_girls = 14
        print(f"Filling GIRLS row ({row_girls}): {girls_data['count']} students")
        print(f"  Setting D{row_girls} (GIRLS) = {girls_data['count']}")
        print(f"  Setting E{row_girls} (TOTAL) = {girls_data['count']}")
        print(f"  Setting F{row_girls} (AB) = {girls_data['ab_count']}")
        sheet.cell(row=row_girls, column=4).value = girls_data['count']  # GIRLS column (D)
        sheet.cell(row=row_girls, column=5).value = girls_data['count']  # TOTAL column (E)
        sheet.cell(row=row_girls, column=6).value = girls_data['ab_count']  # AB column (F)
        
        for grade, col in col_map.items():
            count = girls_data['grade_dist'].get(grade, 0)
            if count > 0:
                sheet.cell(row=row_girls, column=col).value = count
        
        if girls_data['mean_points']:
            sheet.cell(row=row_girls, column=23).value = round(girls_data['mean_points'], 2)
            print(f"  Setting W{row_girls} (Mean Points) = {round(girls_data['mean_points'], 2)}")
        print(f"  ✓ GIRLS row filled successfully")
    else:
        print("No girls data available - GIRLS row not filled")
    
    # Fill TOTAL row (15) with overall statistics
    row_total = 15
    sheet.cell(row=row_total, column=5).value = total_students  # TOTAL
    sheet.cell(row=row_total, column=6).value = ab_count  # AB
    
    for grade, col in col_map.items():
        sheet.cell(row=row_total, column=col).value = grade_dist.get(grade, 0)
    
    if mean_points:
        sheet.cell(row=row_total, column=23).value = round(mean_points, 2)
    
    # Save the updated template
    output_file = 'KCSE ANALYSIS TEMPLATE_2025_FILLED.xlsx'
    wb.save(output_file)
    
    print("="*60)
    print("KCSE ANALYSIS TEMPLATE FILLING SUMMARY")
    print("="*60)
    print(f"Output file: {output_file}")
    print()
    print("Data filled:")
    print(f"  - Total Students: {total_students}")
    print(f"  - School: KUBWEYE SECONDARY SCHOOL")
    print(f"  - AB Count: {ab_count}")
    print(f"  - Mean Points (2025): {round(mean_points, 2) if mean_points else 'N/A'}")
    
    if boys_data:
        print()
        print("  - BOYS:")
        print(f"    Total: {boys_data['count']}")
        print(f"    AB Count: {boys_data['ab_count']}")
        print(f"    Mean Points: {round(boys_data['mean_points'], 2) if boys_data['mean_points'] else 'N/A'}")
    
    if girls_data:
        print()
        print("  - GIRLS:")
        print(f"    Total: {girls_data['count']}")
        print(f"    AB Count: {girls_data['ab_count']}")
        print(f"    Mean Points: {round(girls_data['mean_points'], 2) if girls_data['mean_points'] else 'N/A'}")
    
    print()
    print("Grade Distribution (Overall):")
    for grade in ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']:
        count = grade_dist.get(grade, 0)
        if count > 0:
            print(f"  - {grade}: {count}")
    print()
    print(f"Template saved successfully!")
    print("="*60)

if __name__ == "__main__":
    fill_analysis_template()
