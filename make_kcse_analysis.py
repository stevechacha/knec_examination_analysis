#!/usr/bin/env python3
"""Create complete KCSE Analysis Excel file."""

import pandas as pd
from openpyxl import Workbook
import numpy as np

# Grade to points
gtp = {'A': 12, 'A-': 11, 'B+': 10, 'B': 9, 'B-': 8, 'C+': 7, 'C': 6, 'C-': 5, 'D+': 4, 'D': 3, 'D-': 2, 'E': 1}

def grade_dist(grades):
    g = grades.dropna().astype(str)
    g = g[g != 'nan']
    d = {}
    for gr in ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']:
        d[gr] = len(g[g == gr])
    ab = sum(d.get(x, 0) for x in ['A', 'A-', 'B+', 'B', 'B-'])
    return d, ab

def mean_pts(grades):
    g = grades.dropna().astype(str)
    g = g[g != 'nan']
    pts = [gtp[x] for x in g if x in gtp]
    return np.mean(pts) if pts else None

# Load data
df = pd.read_excel('Students Upload KCSE Results - Template_updated.xlsx')
total = len(df)

# Gender
gender_col = None
for col in df.columns:
    if 'GENDER' in str(col).upper():
        gender_col = col
        break

if gender_col:
    def norm_g(v):
        if pd.isna(v):
            return None
        v = str(v).strip().upper()
        return 'MALE' if v in ['M', 'MALE'] else 'FEMALE' if v in ['F', 'FEMALE'] else None
    df['G'] = df[gender_col].apply(norm_g)
    boys = df[df['G'] == 'MALE'].copy()
    girls = df[df['G'] == 'FEMALE'].copy()
else:
    boys = pd.DataFrame()
    girls = pd.DataFrame()

# Overall
mg = df['MEAN_GRADE'].dropna()
gd, ab = grade_dist(mg)
mp = mean_pts(mg)

# Gender stats
bd = None
if len(boys) > 0:
    bg = boys['MEAN_GRADE'].dropna()
    bd, bab = grade_dist(bg)
    bmp = mean_pts(bg)
    bd = {'count': len(boys), 'dist': bd, 'ab': bab, 'mean': bmp}

gd_g = None
if len(girls) > 0:
    gg = girls['MEAN_GRADE'].dropna()
    gd_g, gab = grade_dist(gg)
    gmp = mean_pts(gg)
    gd_g = {'count': len(girls), 'dist': gd_g, 'ab': gab, 'mean': gmp}

# Subjects
subj_map = {
    'ENG': ('English', '101'), 'KIS': ('Kiswahili', '102'), 'MAT': ('Mathematics', '121'),
    'BIO': ('Biology', '231'), 'PHY': ('Physics', '232'), 'CHE': ('Chemistry', '233'),
    'HIS': ('History and Government', '311'), 'GEO': ('Geography', '312'),
    'CRE': ('C.R.E.', '313'), 'AGR': ('Agriculture', '443'),
    'BST': ('Business Studies', '565'), 'COM': ('Computer Studies', '451')
}

subj_stats = {}
for s, (n, c) in subj_map.items():
    if s in df.columns:
        sg = df[s].dropna().astype(str)
        sg = sg[sg != 'nan']
        if len(sg) > 0:
            sd, sab = grade_dist(sg)
            smp = mean_pts(sg)
            subj_stats[s] = {'name': n, 'code': c, 'total': len(sg), 'dist': sd, 'ab': sab, 'mean': smp}

boys_subj = {}
if len(boys) > 0:
    for s, (n, c) in subj_map.items():
        if s in boys.columns:
            sg = boys[s].dropna().astype(str)
            sg = sg[sg != 'nan']
            if len(sg) > 0:
                sd, sab = grade_dist(sg)
                smp = mean_pts(sg)
                boys_subj[s] = {'name': n, 'code': c, 'count': len(sg), 'dist': sd, 'ab': sab, 'mean': smp}

girls_subj = {}
if len(girls) > 0:
    for s, (n, c) in subj_map.items():
        if s in girls.columns:
            sg = girls[s].dropna().astype(str)
            sg = sg[sg != 'nan']
            if len(sg) > 0:
                sd, sab = grade_dist(sg)
                smp = mean_pts(sg)
                girls_subj[s] = {'name': n, 'code': c, 'count': len(sg), 'dist': sd, 'ab': sab, 'mean': smp}

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "KCSE Analysis 2025"

# Column mapping
cm = {'A': 8, 'A-': 9, 'B+': 10, 'B': 11, 'B-': 12, 'C+': 13, 'C': 14, 'C-': 15, 'D+': 16, 'D': 17, 'D-': 18, 'E': 19}
AB = 7
Y2025 = 25

# ORDER OF MERIT
ws.cell(3, 1).value = "ORDER OF MERIT"
ws.cell(6, 1).value = "SUB COUNTY"
ws.cell(6, 2).value = "SCHOOL"
ws.cell(6, 3).value = "BOYS"
ws.cell(6, 4).value = "GIRLS"
ws.cell(6, 5).value = "TOTAL"
ws.cell(6, 7).value = "AB"
for i, g in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    ws.cell(6, i).value = g
ws.cell(6, 25).value = "2025"

ws.cell(7, 1).value = "MABERA"
ws.cell(7, 2).value = "KUBWEYE SECONDARY SCHOOL"
if bd:
    ws.cell(7, 3).value = bd['count']
if gd_g:
    ws.cell(7, 4).value = gd_g['count']
ws.cell(7, 5).value = total
ws.cell(7, AB).value = ab
for g, c in cm.items():
    ws.cell(7, c).value = gd.get(g, 0)
if mp:
    ws.cell(7, Y2025).value = round(mp, 2)

# PERFORMANCE BY GENDER
ws.cell(9, 1).value = "PERFORMANCE BY GENDER"
ws.cell(12, 1).value = "SCHOOL"
ws.cell(12, 2).value = "GENDER"
ws.cell(12, 3).value = "BOYS"
ws.cell(12, 4).value = "GIRLS"
ws.cell(12, 5).value = "TOTAL"
ws.cell(12, 7).value = "AB"
for i, g in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    ws.cell(12, i).value = g
ws.cell(12, 25).value = "2025"

if bd:
    ws.cell(13, 1).value = "KUBWEYE SECONDARY SCHOOL"
    ws.cell(13, 2).value = "BOYS"
    ws.cell(13, 3).value = bd['count']
    ws.cell(13, 5).value = bd['count']
    ws.cell(13, AB).value = bd['ab']
    for g, c in cm.items():
        ws.cell(13, c).value = bd['dist'].get(g, 0)
    if bd['mean']:
        ws.cell(13, Y2025).value = round(bd['mean'], 2)

if gd_g:
    ws.cell(14, 1).value = "KUBWEYE SECONDARY SCHOOL"
    ws.cell(14, 2).value = "GIRLS"
    ws.cell(14, 4).value = gd_g['count']
    ws.cell(14, 5).value = gd_g['count']
    ws.cell(14, AB).value = gd_g['ab']
    for g, c in cm.items():
        ws.cell(14, c).value = gd_g['dist'].get(g, 0)
    if gd_g['mean']:
        ws.cell(14, Y2025).value = round(gd_g['mean'], 2)

ws.cell(15, 2).value = "TOTAL"
ws.cell(15, 5).value = total
ws.cell(15, AB).value = ab
for g, c in cm.items():
    ws.cell(15, c).value = gd.get(g, 0)
if mp:
    ws.cell(15, Y2025).value = round(mp, 2)

# GENERAL SUBJECT ANALYSIS
ws.cell(17, 1).value = "SUBJECT ANALYSIS BY IMPROVEMENT INDEX"
ws.cell(20, 1).value = "SUBJECT"
ws.cell(20, 2).value = "CODE"
ws.cell(20, 3).value = "BOYS"
ws.cell(20, 4).value = "GIRLS"
ws.cell(20, 5).value = "TOTAL"
ws.cell(20, 7).value = "AB"
for i, g in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    ws.cell(20, i).value = g
ws.cell(20, 25).value = "2025"

r = 40
for s, (n, c) in subj_map.items():
    if s in subj_stats:
        st = subj_stats[s]
        ws.cell(r, 2).value = st['name']
        ws.cell(r, 3).value = st['code']
        ws.cell(r, 6).value = st['total']
        ws.cell(r, AB).value = st['ab']
        for g, c in cm.items():
            ws.cell(r, c).value = st['dist'].get(g, 0)
        if st['mean']:
            ws.cell(r, Y2025).value = round(st['mean'], 2)
        r += 1

# SUBJECT ANALYSIS (BOYS)
ws.cell(36, 1).value = "GENERAL SUBJECT ANALYSIS"
ws.cell(39, 1).value = "SUBJECT"
ws.cell(39, 2).value = "CODE"
ws.cell(39, 3).value = "BOYS"
ws.cell(39, 4).value = "GIRLS"
ws.cell(39, 5).value = "TOTAL"
ws.cell(39, 7).value = "AB"
for i, g in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    ws.cell(39, i).value = g
ws.cell(39, 25).value = "2025"

r = 56
for s, (n, c) in subj_map.items():
    if s in boys_subj:
        st = boys_subj[s]
        ws.cell(r, 2).value = st['name']
        ws.cell(r, 3).value = st['code']
        ws.cell(r, 4).value = st['count']
        ws.cell(r, 6).value = st['count']
        ws.cell(r, AB).value = st['ab']
        for g, c in cm.items():
            ws.cell(r, c).value = st['dist'].get(g, 0)
        if st['mean']:
            ws.cell(r, Y2025).value = round(st['mean'], 2)
        r += 1

# SUBJECT ANALYSIS (GIRLS)
ws.cell(52, 1).value = "SUBJECT ANALYSIS (BOYS)"
ws.cell(55, 1).value = "SUBJECT"
ws.cell(55, 2).value = "CODE"
ws.cell(55, 3).value = "BOYS"
ws.cell(55, 4).value = "GIRLS"
ws.cell(55, 5).value = "TOTAL"
ws.cell(55, 7).value = "AB"
for i, g in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    ws.cell(55, i).value = g
ws.cell(55, 25).value = "2025"

ws.cell(70, 1).value = "SUBJECT ANALYSIS (GIRLS)"
ws.cell(73, 1).value = "SUBJECT"
ws.cell(73, 2).value = "CODE"
ws.cell(73, 3).value = "BOYS"
ws.cell(73, 4).value = "GIRLS"
ws.cell(73, 5).value = "TOTAL"
ws.cell(73, 7).value = "AB"
for i, g in enumerate(['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E'], start=8):
    ws.cell(73, i).value = g
ws.cell(73, 25).value = "2025"

r = 74
for s, (n, c) in subj_map.items():
    if s in girls_subj:
        st = girls_subj[s]
        ws.cell(r, 2).value = st['name']
        ws.cell(r, 3).value = st['code']
        ws.cell(r, 5).value = st['count']
        ws.cell(r, 6).value = st['count']
        ws.cell(r, AB).value = st['ab']
        for g, c in cm.items():
            ws.cell(r, c).value = st['dist'].get(g, 0)
        if st['mean']:
            ws.cell(r, Y2025).value = round(st['mean'], 2)
        r += 1

# TOP 10 BOYS
ws.cell(88, 1).value = "TOP 10 BOYS"
ws.cell(89, 1).value = "INDEX_NO"
ws.cell(89, 2).value = "NAME"
for i, s in enumerate(['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'HIS', 'GEO', 'CRE', 'AGR', 'BST'], start=3):
    ws.cell(89, i).value = s
ws.cell(89, 18).value = "GR"

if len(boys) > 0:
    boys['PTS'] = boys['MEAN_GRADE'].apply(lambda x: gtp.get(str(x).strip(), 0) if pd.notna(x) else 0)
    top_b = boys.nlargest(10, 'PTS')
    r = 90
    for idx, s in top_b.iterrows():
        ws.cell(r, 1).value = str(s['INDEXNO'])
        ws.cell(r, 2).value = s['NAME']
        for i, subj in enumerate(['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'HIS', 'GEO', 'CRE', 'AGR', 'BST'], start=3):
            if subj in s and pd.notna(s[subj]) and str(s[subj]) != 'nan':
                ws.cell(r, i).value = str(s[subj])
        ws.cell(r, 18).value = s['MEAN_GRADE']
        r += 1

# TOP 10 GIRLS
ws.cell(103, 1).value = "TOP 10 GIRLS"
ws.cell(104, 1).value = "INDEX_NO"
ws.cell(104, 2).value = "NAME"
for i, s in enumerate(['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'HIS', 'GEO', 'CRE', 'AGR', 'BST'], start=3):
    ws.cell(104, i).value = s
ws.cell(104, 18).value = "GR"

if len(girls) > 0:
    girls['PTS'] = girls['MEAN_GRADE'].apply(lambda x: gtp.get(str(x).strip(), 0) if pd.notna(x) else 0)
    top_g = girls.nlargest(10, 'PTS')
    r = 105
    for idx, s in top_g.iterrows():
        ws.cell(r, 1).value = str(s['INDEXNO'])
        ws.cell(r, 2).value = s['NAME']
        for i, subj in enumerate(['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'HIS', 'GEO', 'CRE', 'AGR', 'BST'], start=3):
            if subj in s and pd.notna(s[subj]) and str(s[subj]) != 'nan':
                ws.cell(r, i).value = str(s[subj])
        ws.cell(r, 18).value = s['MEAN_GRADE']
        r += 1

# Save
wb.save('KCSE_ANALYSIS_2025_COMPLETE.xlsx')
print("="*70)
print("SUCCESS! File created: KCSE_ANALYSIS_2025_COMPLETE.xlsx")
print("="*70)
print(f"Total Students: {total}")
if bd:
    print(f"Boys: {bd['count']} (AB: {bd['ab']})")
if gd_g:
    print(f"Girls: {gd_g['count']} (AB: {gd_g['ab']})")
print(f"Subjects: {len(subj_stats)}")
print("="*70)
