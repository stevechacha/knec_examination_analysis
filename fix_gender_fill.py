#!/usr/bin/env python3
"""
Fix gender fill - ensure it works correctly.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

print("="*70)
print("FILLING GENDER SECTIONS IN KCSE ANALYSIS TEMPLATE")
print("="*70)

# Grade to points mapping
grade_to_points = {
    'A': 12, 'A-': 11, 'B+': 10, 'B': 9, 'B-': 8,
    'C+': 7, 'C': 6, 'C-': 5, 'D+': 4, 'D': 3, 'D-': 2, 'E': 1
}

def calculate_grade_distribution(mean_grades):
    """Calculate grade distribution."""
    grades = mean_grades.dropna().astype(str)
    distribution = {}
    grade_order = ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']
    for grade in grade_order:
        distribution[grade] = len(grades[grades == grade])
    ab_count = distribution.get('A', 0) + distribution.get('A-', 0) + \
               distribution.get('B+', 0) + distribution.get('B', 0) + distribution.get('B-', 0)
    return distribution, ab_count

def calculate_mean_points(mean_grades):
    """Calculate mean points."""
    grades = mean_grades.dropna().astype(str)
    points_list = [grade_to_points[g] for g in grades if g in grade_to_points]
    return np.mean(points_list) if points_list else None

# Step 1: Load results
print("\nStep 1: Loading results file...")
results_df = pd.read_excel('Students Upload KCSE Results - Template_updated.xlsx')
print(f"   ✓ Loaded {len(results_df)} students")

# Step 2: Find GENDER column
print("\nStep 2: Finding GENDER column...")
gender_col = None
for col in results_df.columns:
    if 'GENDER' in str(col).upper():
        gender_col = col
        break

if not gender_col:
    print("   ✗ ERROR: GENDER column not found!")
    print(f"   Available columns: {list(results_df.columns)}")
    exit(1)

print(f"   ✓ Found GENDER column: '{gender_col}'")

# Step 3: Normalize gender
print("\nStep 3: Processing gender data...")
def normalize_gender(val):
    if pd.isna(val):
        return None
    val_str = str(val).strip().upper()
    if val_str in ['M', 'MALE', 'BOY', 'BOYS']:
        return 'MALE'
    elif val_str in ['F', 'FEMALE', 'GIRL', 'GIRLS']:
        return 'FEMALE'
    return None

results_df['GENDER_NORM'] = results_df[gender_col].apply(normalize_gender)
gender_counts = results_df['GENDER_NORM'].value_counts()
print(f"   Gender distribution: {gender_counts.to_dict()}")

# Step 4: Calculate statistics
print("\nStep 4: Calculating statistics...")
boys_df = results_df[results_df['GENDER_NORM'] == 'MALE']
girls_df = results_df[results_df['GENDER_NORM'] == 'FEMALE']

boys_data = None
if len(boys_df) > 0:
    boys_grades = boys_df['MEAN_GRADE'].dropna()
    boys_dist, boys_ab = calculate_grade_distribution(boys_grades)
    boys_mean = calculate_mean_points(boys_grades)
    boys_data = {
        'count': len(boys_df),
        'grade_dist': boys_dist,
        'ab_count': boys_ab,
        'mean_points': boys_mean
    }
    print(f"   ✓ Boys: {boys_data['count']} students, AB: {boys_data['ab_count']}, Mean: {boys_data['mean_points']:.2f}" if boys_mean else f"   ✓ Boys: {boys_data['count']} students, AB: {boys_data['ab_count']}")

girls_data = None
if len(girls_df) > 0:
    girls_grades = girls_df['MEAN_GRADE'].dropna()
    girls_dist, girls_ab = calculate_grade_distribution(girls_grades)
    girls_mean = calculate_mean_points(girls_grades)
    girls_data = {
        'count': len(girls_df),
        'grade_dist': girls_dist,
        'ab_count': girls_ab,
        'mean_points': girls_mean
    }
    print(f"   ✓ Girls: {girls_data['count']} students, AB: {girls_data['ab_count']}, Mean: {girls_data['mean_points']:.2f}" if girls_mean else f"   ✓ Girls: {girls_data['count']} students, AB: {girls_data['ab_count']}")

# Step 5: Load template
print("\nStep 5: Loading template...")
wb = load_workbook('KCSE ANALYSIS TEMPLATE.xlsx')
sheet = wb.active
print("   ✓ Template loaded")

# Grade column mapping
col_map = {
    'A': 7, 'A-': 8, 'B+': 9, 'B': 10, 'B-': 11,
    'C+': 12, 'C': 13, 'C-': 14, 'D+': 15, 'D': 16, 'D-': 17, 'E': 18
}

# Step 6: Fill Row 7 (ORDER OF MERIT)
print("\nStep 6: Filling Row 7 (ORDER OF MERIT)...")
if boys_data:
    sheet.cell(row=7, column=3).value = boys_data['count']
    print(f"   ✓ Set C7 (BOYS) = {boys_data['count']}")
if girls_data:
    sheet.cell(row=7, column=4).value = girls_data['count']
    print(f"   ✓ Set D7 (GIRLS) = {girls_data['count']}")

# Step 7: Fill Row 13 (BOYS)
print("\nStep 7: Filling Row 13 (BOYS)...")
if boys_data:
    sheet.cell(row=13, column=3).value = boys_data['count']  # C13 - BOYS
    sheet.cell(row=13, column=5).value = boys_data['count']  # E13 - TOTAL
    sheet.cell(row=13, column=6).value = boys_data['ab_count']  # F13 - AB
    
    for grade, col in col_map.items():
        count = boys_data['grade_dist'].get(grade, 0)
        sheet.cell(row=13, column=col).value = count
    
    if boys_data['mean_points']:
        sheet.cell(row=13, column=23).value = round(boys_data['mean_points'], 2)  # W13 - Mean Points
    
    print(f"   ✓ Filled Row 13 with boys data")
    print(f"      C13 (BOYS): {boys_data['count']}")
    print(f"      E13 (TOTAL): {boys_data['count']}")
    print(f"      F13 (AB): {boys_data['ab_count']}")

# Step 8: Fill Row 14 (GIRLS)
print("\nStep 8: Filling Row 14 (GIRLS)...")
if girls_data:
    sheet.cell(row=14, column=4).value = girls_data['count']  # D14 - GIRLS
    sheet.cell(row=14, column=5).value = girls_data['count']  # E14 - TOTAL
    sheet.cell(row=14, column=6).value = girls_data['ab_count']  # F14 - AB
    
    for grade, col in col_map.items():
        count = girls_data['grade_dist'].get(grade, 0)
        sheet.cell(row=14, column=col).value = count
    
    if girls_data['mean_points']:
        sheet.cell(row=14, column=23).value = round(girls_data['mean_points'], 2)  # W14 - Mean Points
    
    print(f"   ✓ Filled Row 14 with girls data")
    print(f"      D14 (GIRLS): {girls_data['count']}")
    print(f"      E14 (TOTAL): {girls_data['count']}")
    print(f"      F14 (AB): {girls_data['ab_count']}")

# Step 9: Save
print("\nStep 9: Saving file...")
output_file = 'KCSE ANALYSIS TEMPLATE_2025_FILLED.xlsx'
wb.save(output_file)
print(f"   ✓ Saved to: {output_file}")

# Step 10: Verify
print("\nStep 10: Verification...")
row13_c = sheet.cell(13, 3).value
row14_d = sheet.cell(14, 4).value
row7_c = sheet.cell(7, 3).value
row7_d = sheet.cell(7, 4).value

print(f"   Row 7, Column C (BOYS): {row7_c}")
print(f"   Row 7, Column D (GIRLS): {row7_d}")
print(f"   Row 13, Column C (BOYS): {row13_c}")
print(f"   Row 14, Column D (GIRLS): {row14_d}")

if row13_c is not None or row14_d is not None:
    print("\n" + "="*70)
    print("✓ SUCCESS: Gender sections have been filled!")
    print("="*70)
else:
    print("\n" + "="*70)
    print("✗ WARNING: Gender sections may not have been filled correctly")
    print("="*70)

print()
