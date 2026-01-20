"""
Excel Processing Module
=======================
Handles Excel template processing and data population.
"""

import logging
import re
from pathlib import Path
from typing import Dict, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelProcessor:
    """Process Excel templates and populate with extracted data."""
    
    def __init__(self, logger: logging.Logger, config=None):
        """
        Initialize Excel processor.
        
        Args:
            logger: Logger instance
            config: Optional configuration object
        """
        self.logger = logger
        self.config = config
        
        self.index_column = 'INDEXNO'
        self.name_column = 'NAME'
        
        if config:
            if hasattr(config, 'index_column'):
                self.index_column = config.index_column
            if hasattr(config, 'name_column'):
                self.name_column = config.name_column
    
    def normalize_index_number(self, index_num) -> Optional[str]:
        """Normalize index number format."""
        if pd.isna(index_num):
            return None
        
        index_str = str(index_num).strip()
        if index_str.endswith('.0'):
            index_str = index_str[:-2]
        index_str = re.sub(r'/\d{4}$', '', index_str)
        index_str = re.sub(r'\s+', '', index_str)
        
        return index_str if index_str else None
    
    def find_column(self, df: pd.DataFrame, possible_names: list) -> Optional[str]:
        """Find column by possible names."""
        for col in df.columns:
            col_lower = str(col).lower()
            for possible in possible_names:
                if possible.lower() in col_lower:
                    return col
        return None
    
    def update_template(self, template_path: Path, extracted_data: Dict, 
                       output_path: Path) -> Path:
        """
        Update Excel template with extracted data.
        
        Args:
            template_path: Path to Excel template
            extracted_data: Dictionary of extracted data keyed by index number
            output_path: Path for output file
            
        Returns:
            Path to output file
        """
        self.logger.info(f"Loading template: {template_path}")
        
        # Read Excel file
        df = pd.read_excel(template_path)
        self.logger.info(f"Template loaded: {len(df)} rows, {len(df.columns)} columns")
        
        # Find index column
        index_col = self.find_column(df, ['index number', 'index no', 'index', 'indexno', 'index_number'])
        if not index_col:
            raise ValueError("Could not find index number column in template")
        
        self.logger.info(f"Using index column: {index_col}")
        
        # Find subject columns
        subject_columns = [
            'ENG', 'KIS', 'MAT', 'BIO', 'CHE', 'GEO', 'CRE', 'AGR', 'BST',
            'HIS', 'PHY', 'COM', 'IRE', 'HRE', 'KSL', 'FAS', 'LIT', 'GSC',
            'AD', 'HSC', 'MEAN_GRADE', 'MEAN GRADE'
        ]
        
        available_subject_cols = {col: col for col in subject_columns if col in df.columns}
        
        # Find mean grade column
        mean_grade_col = None
        for col in ['MEAN_GRADE', 'mean_grade', 'MEAN GRADE', 'Mean Grade']:
            if col in df.columns:
                mean_grade_col = col
                break
        
        self.logger.info(f"Found {len(available_subject_cols)} subject columns")
        if mean_grade_col:
            self.logger.info(f"Mean grade column: {mean_grade_col}")
        
        # Create mapping of normalized index numbers to row indices
        index_to_row = {}
        for idx, row in df.iterrows():
            index_num = self.normalize_index_number(row[index_col])
            if index_num:
                index_to_row[index_num] = idx
                # Also create partial matches (last 9-10 digits)
                if len(index_num) >= 9:
                    partial_key = index_num[-9:]
                    if partial_key not in index_to_row:
                        index_to_row[partial_key] = idx
                    if len(index_num) >= 10:
                        partial_key = index_num[-10:]
                        if partial_key not in index_to_row:
                            index_to_row[partial_key] = index_to_row.get(index_num[-9:], idx)
        
        # Normalize extracted data keys
        normalized_extracted = {}
        for key, value in extracted_data.items():
            normalized_key = self.normalize_index_number(key)
            if normalized_key:
                normalized_extracted[normalized_key] = value
        
        # Update grades
        updated_count = 0
        for normalized_index, data in normalized_extracted.items():
            # Try exact match first
            row_idx = index_to_row.get(normalized_index)
            
            # Try partial match if exact match failed
            if row_idx is None and len(normalized_index) >= 9:
                partial_key = normalized_index[-9:]
                row_idx = index_to_row.get(partial_key)
                if row_idx is None and len(normalized_index) >= 10:
                    partial_key = normalized_index[-10:]
                    row_idx = index_to_row.get(partial_key)
            
            if row_idx is not None:
                subject_grades = data.get('subject_grades', {})
                mean_grade = data.get('mean_grade')
                
                # Update subject columns
                updated_subjects = []
                for subject_col, grade in subject_grades.items():
                    if subject_col in available_subject_cols:
                        col_name = available_subject_cols[subject_col]
                        if df[col_name].dtype != 'object':
                            df[col_name] = df[col_name].astype(str)
                        df.at[row_idx, col_name] = str(grade)
                        updated_subjects.append(f"{subject_col}:{grade}")
                
                # Update mean grade
                if mean_grade_col and mean_grade:
                    if df[mean_grade_col].dtype != 'object':
                        df[mean_grade_col] = df[mean_grade_col].astype(str)
                    df.at[row_idx, mean_grade_col] = str(mean_grade)
                
                if updated_subjects:
                    excel_index = df.at[row_idx, index_col]
                    self.logger.debug(
                        f"Updated row {row_idx+1} (Index: {excel_index}) "
                        f"with {len(updated_subjects)} subjects"
                    )
                    updated_count += 1
            else:
                self.logger.warning(f"Index {normalized_index} not found in template")
        
        # Save the updated Excel file
        self.logger.info(f"Saving to: {output_path}")
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Set index number column format to text to prevent .0 suffix
        try:
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Find index column
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == index_col:
                    col_letter = get_column_letter(col_idx)
                    # Set entire column to text format
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row_idx}']
                        if cell.value:
                            val = str(cell.value).replace('.0', '')
                            cell.value = val
                            cell.number_format = '@'  # Text format
                    break
            
            wb.save(output_path)
            wb.close()
        except Exception as e:
            self.logger.warning(f"Could not set text format for index column: {e}")
        
        self.logger.info(f"Successfully updated {updated_count} records")
        return output_path
