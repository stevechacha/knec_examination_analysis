#!/usr/bin/env python3
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

log = []
def log_msg(msg):
    log.append(msg)
    print(msg)

log_msg("Starting gender fill process...")

# Load data
df = pd.read_excel('Students Upload KCSE Results - Template_updated.xlsx')
log_msg(f"Loaded {len(df)} students")

# Find GENDER column
gender_col = None
for col in df.columns:
    if 'GENDER' in str(col).upper():
        gender_col = col
        break

if not gender_col:
    log_msg("ERROR: GENDER column not found!")
    with open('gender_fill_log.txt', 'w') as f:
        f.write('\n'.join(log))
    exit(1)

log_msg(f"Found GENDER column: {gender_col}")

# Normalize gender
def norm(v):
    if pd.isna(v):
        return None
    v = str(v).strip().upper()
    return 'MALE' if v in ['M', 'MALE'] else 'FEMALE' if v in ['F', 'FEMALE'] else None

df['GEND'] = df[gender_col].apply(norm)
log_msg(f"Gender distribution: {df['GEND'].value_counts().to_dict()}")

# Calculate stats
boys = df[df['GEND'] == 'MALE']
girls = df[df['GEND'] == 'FEMALE']
log_msg(f"Boys: {len(boys)}, Girls: {len(girls)}")

# Grade to points
gtp = {'A': 12, 'A-': 11, 'B+': 10, 'B': 9, 'B-': 8, 'C+': 7, 'C': 6, 'C-': 5, 'D+': 4, 'D': 3, 'D-': 2, 'E': 1}

def grade_dist(grades):
    g = grades.dropna().astype(str)
    d = {}
    for gr in ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']:
        d[gr] = len(g[g == gr])
    ab = sum(d.get(x, 0) for x in ['A', 'A-', 'B+', 'B', 'B-'])
    return d, ab

def mean_pts(grades):
    g = grades.dropna().astype(str)
    pts = [gtp[x] for x in g if x in gtp]
    return np.mean(pts) if pts else None

# Boys stats
if len(boys) > 0:
    bg = boys['MEAN_GRADE'].dropna()
    bd, bab = grade_dist(bg)
    bmp = mean_pts(bg)
    log_msg(f"Boys stats: count={len(boys)}, AB={bab}, mean={bmp:.2f if bmp else None}")
else:
    bd, bab, bmp = None, 0, None

# Girls stats
if len(girls) > 0:
    gg = girls['MEAN_GRADE'].dropna()
    gd, gab = grade_dist(gg)
    gmp = mean_pts(gg)
    log_msg(f"Girls stats: count={len(girls)}, AB={gab}, mean={gmp:.2f if gmp else None}")
else:
    gd, gab, gmp = None, 0, None

# Load template (always from original)
wb = load_workbook('KCSE ANALYSIS TEMPLATE.xlsx')
sheet = wb.active
log_msg("Loaded template")

# Column mapping
cm = {'A': 7, 'A-': 8, 'B+': 9, 'B': 10, 'B-': 11, 'C+': 12, 'C': 13, 'C-': 14, 'D+': 15, 'D': 16, 'D-': 17, 'E': 18}

# Fill Row 7
if len(boys) > 0:
    sheet.cell(7, 3).value = len(boys)
    log_msg(f"Set C7 = {len(boys)}")
if len(girls) > 0:
    sheet.cell(7, 4).value = len(girls)
    log_msg(f"Set D7 = {len(girls)}")

# Fill Row 13 (BOYS)
if len(boys) > 0 and bd:
    sheet.cell(13, 3).value = len(boys)  # C13
    sheet.cell(13, 5).value = len(boys)  # E13
    sheet.cell(13, 6).value = bab        # F13
    for grade, col in cm.items():
        sheet.cell(13, col).value = bd.get(grade, 0)
    if bmp:
        sheet.cell(13, 23).value = round(bmp, 2)
    log_msg(f"Filled Row 13 - C13={len(boys)}, E13={len(boys)}, F13={bab}")

# Fill Row 14 (GIRLS)
if len(girls) > 0 and gd:
    sheet.cell(14, 4).value = len(girls)  # D14
    sheet.cell(14, 5).value = len(girls)  # E14
    sheet.cell(14, 6).value = gab         # F14
    for grade, col in cm.items():
        sheet.cell(14, col).value = gd.get(grade, 0)
    if gmp:
        sheet.cell(14, 23).value = round(gmp, 2)
    log_msg(f"Filled Row 14 - D14={len(girls)}, E14={len(girls)}, F14={gab}")

# Save
wb.save('KCSE ANALYSIS TEMPLATE_2025_FILLED.xlsx')
log_msg("Saved to KCSE ANALYSIS TEMPLATE_2025_FILLED.xlsx")

# Verify
log_msg("\nVerification:")
log_msg(f"  C7 (BOYS): {sheet.cell(7, 3).value}")
log_msg(f"  D7 (GIRLS): {sheet.cell(7, 4).value}")
log_msg(f"  C13 (BOYS): {sheet.cell(13, 3).value}")
log_msg(f"  D14 (GIRLS): {sheet.cell(14, 4).value}")

# Write log
with open('gender_fill_log.txt', 'w') as f:
    f.write('\n'.join(log))

log_msg("\nDone! Check gender_fill_log.txt for details.")
