#!/usr/bin/env python3
"""
Script to fix duplicate index numbers in Upload Student Profiles by using KCSE results
to assign the correct unique index number to each student.
"""

import pandas as pd
from pathlib import Path
import re
from itertools import permutations

def normalize_name(name):
    """Normalize name for matching."""
    if pd.isna(name):
        return None
    name_str = str(name).strip().upper()
    name_str = re.sub(r'\s+', ' ', name_str)
    return name_str

def find_name_column(df, possible_names):
    """Find the name column in a dataframe."""
    for col in df.columns:
        col_lower = str(col).lower()
        for possible in possible_names:
            if possible.lower() in col_lower:
                return col
    return None

def find_best_match(profile_name, name_to_index):
    """Try to find best matching name with fuzzy matching."""
    if not profile_name:
        return None
    
    profile_parts = [p.strip() for p in profile_name.split() if p.strip()]
    
    # Exact match first
    if profile_name in name_to_index:
        return name_to_index[profile_name]
    
    if len(profile_parts) >= 2:
        # Try all permutations
        for perm in permutations(profile_parts):
            perm_name = ' '.join(perm)
            if perm_name in name_to_index:
                return name_to_index[perm_name]
        
        # Try reversed order
        if len(profile_parts) == 2:
            reversed_name = f"{profile_parts[1]} {profile_parts[0]}"
            if reversed_name in name_to_index:
                return name_to_index[reversed_name]
        elif len(profile_parts) == 3:
            variations = [
                f"{profile_parts[2]} {profile_parts[0]} {profile_parts[1]}",
                f"{profile_parts[2]} {profile_parts[1]} {profile_parts[0]}",
                f"{profile_parts[0]} {profile_parts[2]} {profile_parts[1]}",
            ]
            for var in variations:
                if var in name_to_index:
                    return name_to_index[var]
        
        # Try last name + first name
        last_first = f"{profile_parts[-1]} {profile_parts[0]}"
        if last_first in name_to_index:
            return name_to_index[last_first]
    
    # Try partial matching
    for kcse_name in name_to_index.keys():
        kcse_parts = set(kcse_name.split())
        significant_parts = [p for p in profile_parts if len(p) > 2]
        if significant_parts and all(any(p in kp for kp in kcse_parts) for p in significant_parts):
            exact_matches = len([p for p in profile_parts if p in kcse_parts])
            if exact_matches >= 2:
                return name_to_index[kcse_name]
    
    # Try fuzzy matching with spelling variations
    for kcse_name in name_to_index.keys():
        kcse_parts_list = kcse_name.split()
        if len(kcse_parts_list) == len(profile_parts):
            matches = 0
            for i, profile_part in enumerate(profile_parts):
                if i < len(kcse_parts_list):
                    kcse_part = kcse_parts_list[i]
                    if profile_part == kcse_part:
                        matches += 1
                    elif len(profile_part) == len(kcse_part) and len(profile_part) > 3:
                        diff_count = sum(1 for a, b in zip(profile_part, kcse_part) if a != b)
                        if diff_count <= 1:
                            matches += 1
                    elif profile_part in kcse_part or kcse_part in profile_part:
                        matches += 1
            
            if matches >= 2:
                return name_to_index[kcse_name]
    
    return None

def fix_duplicate_index_numbers():
    """Fix duplicate index numbers by matching with KCSE results."""
    script_dir = Path(__file__).parent
    
    # File paths
    profiles_file = script_dir / "Upload Student Profiles_filled_complete.xlsx"
    kcse_file = script_dir / "Students Upload KCSE Results - Template_updated.xlsx"
    
    if not profiles_file.exists():
        profiles_file = script_dir / "Upload Student Profiles_filled.xlsx"
    
    if not profiles_file.exists():
        profiles_file = script_dir / "Upload Student Profiles.xlsx"
    
    if not kcse_file.exists():
        kcse_file = script_dir / "Students Upload KCSE Results - Template.xlsx"
    
    if not profiles_file.exists():
        print(f"Error: Profiles file not found")
        return False
    
    if not kcse_file.exists():
        print(f"Error: KCSE file not found")
        return False
    
    print(f"Reading {profiles_file.name}...")
    profiles_df = pd.read_excel(profiles_file)
    print(f"  Rows: {len(profiles_df)}")
    
    print(f"\nReading {kcse_file.name}...")
    kcse_df = pd.read_excel(kcse_file)
    print(f"  Rows: {len(kcse_df)}")
    
    # Find columns
    name_col_profiles = find_name_column(profiles_df, ['name', 'student name', 'full name'])
    index_col_profiles = find_name_column(profiles_df, ['index number', 'index no', 'index'])
    name_col_kcse = find_name_column(kcse_df, ['name', 'student name', 'full name'])
    index_col_kcse = find_name_column(kcse_df, ['index number', 'index no', 'index'])
    
    if not name_col_profiles or not index_col_profiles:
        print("Error: Could not find required columns in profiles file")
        return False
    
    if not name_col_kcse or not index_col_kcse:
        print("Error: Could not find required columns in KCSE file")
        return False
    
    # Create mapping from KCSE names to index numbers
    name_to_index = {}
    index_to_name = {}  # Track which index belongs to which name
    
    for idx, row in kcse_df.iterrows():
        name = normalize_name(row[name_col_kcse])
        index_num = row[index_col_kcse]
        if name and pd.notna(index_num):
            index_str = str(index_num).strip()
            if index_str.endswith('.0'):
                index_str = index_str[:-2]
            name_to_index[name] = index_str
            index_to_name[index_str] = name
    
    print(f"\nCreated mapping with {len(name_to_index)} KCSE entries")
    
    # Convert index column to string
    profiles_df[index_col_profiles] = profiles_df[index_col_profiles].astype(str)
    profiles_df[index_col_profiles] = profiles_df[index_col_profiles].replace('nan', '')
    
    # Find duplicates
    duplicates_mask = profiles_df[index_col_profiles].duplicated(keep=False) & (profiles_df[index_col_profiles] != '')
    duplicates_df = profiles_df[duplicates_mask].copy()
    
    print(f"\nFound {len(duplicates_df)} rows with duplicate index numbers")
    
    # Strategy: Clear all duplicate index numbers, then re-assign properly
    # Track which index numbers are already assigned to avoid duplicates
    assigned_indices = {}  # index -> (row_idx, name)
    
    # First, clear all duplicate index numbers
    duplicate_mask = profiles_df[index_col_profiles].duplicated(keep=False) & (profiles_df[index_col_profiles] != '')
    profiles_df.loc[duplicate_mask, index_col_profiles] = ''
    cleared_count = duplicate_mask.sum()
    print(f"Cleared {cleared_count} duplicate index numbers")
    
    # Now re-match all students, ensuring no duplicates
    fixed_count = 0
    conflicts = []
    
    # Sort by name to process consistently
    for idx, row in profiles_df.iterrows():
        profile_name = normalize_name(row[name_col_profiles])
        current_index = str(row[index_col_profiles]).strip()
        
        if not profile_name:
            continue
        
        # If already has a valid unique index, skip
        if current_index and current_index != '' and current_index not in assigned_indices:
            assigned_indices[current_index] = (idx, profile_name)
            continue
        
        # Find best match
        matched_index = find_best_match(profile_name, name_to_index)
        
        if matched_index:
            # Check if this index is already assigned
            if matched_index in assigned_indices:
                # Conflict - this index is already assigned to another student
                existing_idx, existing_name = assigned_indices[matched_index]
                
                # Determine which student matches better
                profile_exact = profile_name in name_to_index
                existing_exact = existing_name in name_to_index
                
                if profile_exact and not existing_exact:
                    # Current student has exact match, reassign
                    profiles_df.at[existing_idx, index_col_profiles] = ''
                    # Try to find alternative for existing student
                    alt_match = find_best_match(existing_name, {k: v for k, v in name_to_index.items() if v != matched_index})
                    if alt_match and alt_match not in assigned_indices:
                        profiles_df.at[existing_idx, index_col_profiles] = alt_match
                        assigned_indices[alt_match] = (existing_idx, existing_name)
                    assigned_indices[matched_index] = (idx, profile_name)
                    profiles_df.at[idx, index_col_profiles] = matched_index
                    fixed_count += 1
                    conflicts.append(f"{profile_name} (exact) vs {existing_name}")
                elif not profile_exact and existing_exact:
                    # Existing student has exact match, keep it
                    # Try to find alternative for current student
                    alt_match = find_best_match(profile_name, {k: v for k, v in name_to_index.items() if v != matched_index})
                    if alt_match and alt_match not in assigned_indices:
                        profiles_df.at[idx, index_col_profiles] = alt_match
                        assigned_indices[alt_match] = (idx, profile_name)
                        fixed_count += 1
                    else:
                        conflicts.append(f"{profile_name} (fuzzy) vs {existing_name} (exact)")
                else:
                    # Both fuzzy matches or both exact - keep first assignment
                    conflicts.append(f"{profile_name} vs {existing_name} (both fuzzy)")
            else:
                # Index available, assign it
                profiles_df.at[idx, index_col_profiles] = matched_index
                assigned_indices[matched_index] = (idx, profile_name)
                fixed_count += 1
    
    if conflicts:
        print(f"\nConflicts resolved: {len(conflicts)}")
        for conflict in conflicts[:5]:
            print(f"  {conflict}")
        if len(conflicts) > 5:
            print(f"  ... and {len(conflicts) - 5} more")
    
    # Re-check for remaining duplicates
    remaining_duplicates = profiles_df[profiles_df[index_col_profiles].duplicated(keep=False) & (profiles_df[index_col_profiles] != '')]
    
    print(f"\nFixed {fixed_count} duplicate entries")
    print(f"Remaining duplicates: {len(remaining_duplicates)}")
    
    if len(remaining_duplicates) > 0:
        print("\nRemaining duplicate index numbers:")
        for idx, row in remaining_duplicates.iterrows():
            index_val = str(row[index_col_profiles]).strip()
            name = row[name_col_profiles]
            kcse_name = index_to_name.get(index_val, 'Not found in KCSE')
            print(f"  {name} -> Index: {index_val} (KCSE: {kcse_name})")
    
    # Ensure index numbers are stored as strings without .0 suffix
    def clean_index_number(val):
        if pd.isna(val) or val == '' or str(val).lower() == 'nan':
            return ''
        try:
            val_float = float(val)
            val_int = int(val_float)
            return str(val_int)
        except (ValueError, TypeError):
            return str(val).replace('.0', '')
    
    profiles_df[index_col_profiles] = profiles_df[index_col_profiles].apply(clean_index_number)
    
    # Save fixed file
    output_file = profiles_file.parent / f"{profiles_file.stem}_no_duplicates.xlsx"
    profiles_df.to_excel(output_file, index=False, engine='openpyxl')
    
    # Set column format to text using openpyxl to prevent .0 from appearing
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Find INDEX NUMBER column
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == index_col_profiles:
                col_letter = get_column_letter(col_idx)
                # Set entire column to text format
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row_idx}']
                    if cell.value:
                        val = str(cell.value).replace('.0', '')
                        cell.value = val
                        cell.number_format = '@'  # Text format
                break
        
        wb.save(output_file)
        wb.close()
    except ImportError:
        pass  # openpyxl already imported via pandas
    
    print(f"\nSaved fixed file: {output_file.name}")
    print("✓ Index numbers saved as text without .0 suffix")
    
    return True

if __name__ == "__main__":
    fix_duplicate_index_numbers()
