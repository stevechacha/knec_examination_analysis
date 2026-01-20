#!/usr/bin/env python3
"""
Complete KCSE Analysis Template Filler - Fills all sections.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

# Grade to points mapping (KCSE system)
grade_to_points = {
    'A': 12, 'A-': 11, 'B+': 10, 'B': 9, 'B-': 8,
    'C+': 7, 'C': 6, 'C-': 5, 'D+': 4, 'D': 3, 'D-': 2, 'E': 1
}

def calculate_grade_distribution(mean_grades):
    """Calculate grade distribution from mean grades."""
    grades = mean_grades.dropna().astype(str)
    distribution = {}
    grade_order = ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']
    for grade in grade_order:
        distribution[grade] = len(grades[grades == grade])
    ab_count = distribution.get('A', 0) + distribution.get('A-', 0) + \
               distribution.get('B+', 0) + distribution.get('B', 0) + distribution.get('B-', 0)
    return distribution, ab_count

def calculate_mean_points(mean_grades):
    """Calculate mean points from grades."""
    grades = mean_grades.dropna().astype(str)
    points_list = [grade_to_points[g] for g in grades if g in grade_to_points]
    return np.mean(points_list) if points_list else None

def find_gender_column(df):
    """Find the gender column in the dataframe."""
    for col in df.columns:
        if 'GENDER' in str(col).upper():
            return col
    return None

def normalize_gender(val):
    """Normalize gender values."""
    if pd.isna(val):
        return None
    val_str = str(val).strip().upper()
    if val_str in ['M', 'MALE', 'BOY', 'BOYS']:
        return 'MALE'
    elif val_str in ['F', 'FEMALE', 'GIRL', 'GIRLS']:
        return 'FEMALE'
    return None

def calculate_subject_statistics(results_df):
    """Calculate statistics for each subject."""
    subject_cols = ['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'GEO', 'CRE', 'AGR', 'BST', 'COM', 'HIS']
    subject_stats = {}
    
    for subject in subject_cols:
        if subject in results_df.columns:
            grades = results_df[subject].dropna().astype(str)
            # Filter out 'nan' strings
            grades = grades[grades != 'nan']
            if len(grades) > 0:
                dist, ab = calculate_grade_distribution(grades)
                mean_pts = calculate_mean_points(grades)
                subject_stats[subject] = {
                    'count': len(grades),
                    'grade_dist': dist,
                    'ab_count': ab,
                    'mean_points': mean_pts
                }
    
    return subject_stats

print("="*70)
print("FILLING COMPLETE KCSE ANALYSIS TEMPLATE - 2025")
print("="*70)

# Load results
print("\n1. Loading results data...")
results_df = pd.read_excel('Students Upload KCSE Results - Template_updated.xlsx')
total_students = len(results_df)
print(f"   ✓ Loaded {total_students} students")

# Find gender column
gender_col = find_gender_column(results_df)
if gender_col:
    print(f"   ✓ Found gender column: '{gender_col}'")
    results_df['GENDER_NORM'] = results_df[gender_col].apply(normalize_gender)
    has_gender = True
else:
    print("   ⚠ Gender column not found")
    has_gender = False

# Calculate overall statistics
print("\n2. Calculating overall statistics...")
mean_grades = results_df['MEAN_GRADE'].dropna()
grade_dist, ab_count = calculate_grade_distribution(mean_grades)
mean_points = calculate_mean_points(mean_grades)
print(f"   ✓ AB Count: {ab_count}")
print(f"   ✓ Mean Points: {mean_points:.2f}" if mean_points else "   ✓ Mean Points: N/A")

# Calculate gender statistics
boys_data = None
girls_data = None
if has_gender:
    print("\n3. Calculating gender statistics...")
    boys_df = results_df[results_df['GENDER_NORM'] == 'MALE']
    girls_df = results_df[results_df['GENDER_NORM'] == 'FEMALE']
    
    if len(boys_df) > 0:
        bg = boys_df['MEAN_GRADE'].dropna()
        bd, bab = calculate_grade_distribution(bg)
        bmp = calculate_mean_points(bg)
        boys_data = {'count': len(boys_df), 'grade_dist': bd, 'ab_count': bab, 'mean_points': bmp}
        print(f"   ✓ Boys: {boys_data['count']} students")
    
    if len(girls_df) > 0:
        gg = girls_df['MEAN_GRADE'].dropna()
        gd, gab = calculate_grade_distribution(gg)
        gmp = calculate_mean_points(gg)
        girls_data = {'count': len(girls_df), 'grade_dist': gd, 'ab_count': gab, 'mean_points': gmp}
        print(f"   ✓ Girls: {girls_data['count']} students")

# Calculate subject statistics
print("\n4. Calculating subject statistics...")
subject_stats = calculate_subject_statistics(results_df)
print(f"   ✓ Calculated stats for {len(subject_stats)} subjects")

# Load template
print("\n5. Loading template...")
wb = load_workbook('KCSE ANALYSIS TEMPLATE.xlsx')
sheet = wb.active
print("   ✓ Template loaded")

# Grade column mapping
col_map = {
    'A': 7, 'A-': 8, 'B+': 9, 'B': 10, 'B-': 11,
    'C+': 12, 'C': 13, 'C-': 14, 'D+': 15, 'D': 16, 'D-': 17, 'E': 18
}

# Fill ORDER OF MERIT section (Row 7)
print("\n6. Filling ORDER OF MERIT section (Row 7)...")
sheet.cell(7, 2).value = "KUBWEYE SECONDARY SCHOOL"  # SCHOOL
if boys_data:
    sheet.cell(7, 3).value = boys_data['count']  # BOYS
if girls_data:
    sheet.cell(7, 4).value = girls_data['count']  # GIRLS
sheet.cell(7, 5).value = total_students  # TOTAL
sheet.cell(7, 6).value = ab_count  # AB
for grade, col in col_map.items():
    sheet.cell(7, col).value = grade_dist.get(grade, 0)
if mean_points:
    sheet.cell(7, 23).value = round(mean_points, 2)  # 2025 Mean Points
print("   ✓ Row 7 filled")

# Fill PERFORMANCE BY GENDER section
print("\n7. Filling PERFORMANCE BY GENDER section...")

# Row 13 - BOYS
if boys_data:
    sheet.cell(13, 3).value = boys_data['count']  # C13 - BOYS
    sheet.cell(13, 5).value = boys_data['count']  # E13 - TOTAL
    sheet.cell(13, 6).value = boys_data['ab_count']  # F13 - AB
    for grade, col in col_map.items():
        sheet.cell(13, col).value = boys_data['grade_dist'].get(grade, 0)
    if boys_data['mean_points']:
        sheet.cell(13, 23).value = round(boys_data['mean_points'], 2)  # W13
    print("   ✓ Row 13 (BOYS) filled")

# Row 14 - GIRLS
if girls_data:
    sheet.cell(14, 4).value = girls_data['count']  # D14 - GIRLS
    sheet.cell(14, 5).value = girls_data['count']  # E14 - TOTAL
    sheet.cell(14, 6).value = girls_data['ab_count']  # F14 - AB
    for grade, col in col_map.items():
        sheet.cell(14, col).value = girls_data['grade_dist'].get(grade, 0)
    if girls_data['mean_points']:
        sheet.cell(14, 23).value = round(girls_data['mean_points'], 2)  # W14
    print("   ✓ Row 14 (GIRLS) filled")

# Row 15 - TOTAL
sheet.cell(15, 5).value = total_students  # E15 - TOTAL
sheet.cell(15, 6).value = ab_count  # F15 - AB
for grade, col in col_map.items():
    sheet.cell(15, col).value = grade_dist.get(grade, 0)
if mean_points:
    sheet.cell(15, 23).value = round(mean_points, 2)  # W15
print("   ✓ Row 15 (TOTAL) filled")

# Fill SUBJECT ANALYSIS section if it exists (around row 20+)
print("\n8. Filling SUBJECT ANALYSIS section...")
subject_code_map = {
    'ENG': '101', 'KIS': '102', 'MAT': '121', 'BIO': '231', 'PHY': '232', 'CHE': '233',
    'GEO': '312', 'CRE': '313', 'AGR': '443', 'BST': '565', 'COM': '451', 'HIS': '311'
}

subject_row_start = 20  # Adjust based on template structure
row_num = subject_row_start

for subject, code in subject_code_map.items():
    if subject in subject_stats:
        stats = subject_stats[subject]
        # Find the row for this subject (may need adjustment)
        # For now, we'll try to find it or add it
        sheet.cell(row_num, 1).value = subject  # SUBJECT
        sheet.cell(row_num, 2).value = code  # CODE
        sheet.cell(row_num, 5).value = stats['count']  # TOTAL
        sheet.cell(row_num, 6).value = stats['ab_count']  # AB
        for grade, col in col_map.items():
            sheet.cell(row_num, col).value = stats['grade_dist'].get(grade, 0)
        if stats['mean_points']:
            sheet.cell(row_num, 23).value = round(stats['mean_points'], 2)  # 2025
        row_num += 1

print(f"   ✓ Filled {len(subject_stats)} subjects")

# Save
print("\n9. Saving template...")
output_file = 'KCSE ANALYSIS TEMPLATE_2025_FILLED.xlsx'
wb.save(output_file)
print(f"   ✓ Saved to: {output_file}")

# Summary
print("\n" + "="*70)
print("FILLING COMPLETE!")
print("="*70)
print(f"\nSummary:")
print(f"  - Total Students: {total_students}")
print(f"  - School: KUBWEYE SECONDARY SCHOOL")
print(f"  - AB Count: {ab_count}")
print(f"  - Mean Points (2025): {mean_points:.2f}" if mean_points else "  - Mean Points: N/A")
if boys_data:
    print(f"  - Boys: {boys_data['count']} (AB: {boys_data['ab_count']})")
if girls_data:
    print(f"  - Girls: {girls_data['count']} (AB: {girls_data['ab_count']})")
print(f"  - Subjects analyzed: {len(subject_stats)}")
print("\n" + "="*70)
print("Template completely filled!")
print("="*70)
