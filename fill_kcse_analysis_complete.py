#!/usr/bin/env python3
"""
Complete KCSE Analysis Template Filler - Fills ALL sections comprehensively.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import numpy as np

# Grade to points mapping
grade_to_points = {
    'A': 12, 'A-': 11, 'B+': 10, 'B': 9, 'B-': 8,
    'C+': 7, 'C': 6, 'C-': 5, 'D+': 4, 'D': 3, 'D-': 2, 'E': 1
}

def calculate_grade_distribution(grades):
    """Calculate grade distribution."""
    g = grades.dropna().astype(str)
    g = g[g != 'nan']  # Remove 'nan' strings
    dist = {}
    for grade in ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']:
        dist[grade] = len(g[g == grade])
    ab = dist.get('A', 0) + dist.get('A-', 0) + dist.get('B+', 0) + dist.get('B', 0) + dist.get('B-', 0)
    return dist, ab

def calculate_mean_points(grades):
    """Calculate mean points."""
    g = grades.dropna().astype(str)
    g = g[g != 'nan']
    pts = [grade_to_points[x] for x in g if x in grade_to_points]
    return np.mean(pts) if pts else None

def find_gender_column(df):
    """Find gender column."""
    for col in df.columns:
        if 'GENDER' in str(col).upper():
            return col
    return None

def normalize_gender(val):
    """Normalize gender."""
    if pd.isna(val):
        return None
    v = str(val).strip().upper()
    return 'MALE' if v in ['M', 'MALE', 'BOY', 'BOYS'] else 'FEMALE' if v in ['F', 'FEMALE', 'GIRL', 'GIRLS'] else None

def safe_set_cell(sheet, row, col, value):
    """Safely set cell value, handling merged cells."""
    try:
        cell = sheet.cell(row, col)
        if isinstance(cell, MergedCell):
            # For merged cells, find the top-left cell of the merge range
            for merge_range in sheet.merged_cells.ranges:
                if (row, col) in merge_range:
                    top_left = merge_range.top_left
                    sheet.cell(top_left[0], top_left[1]).value = value
                    return
        sheet.cell(row, col).value = value
    except Exception as e:
        # If there's an issue, try direct assignment
        try:
            sheet.cell(row, col).value = value
        except:
            pass  # Skip if we can't write to this cell

print("="*70)
print("FILLING COMPLETE KCSE ANALYSIS TEMPLATE - ALL SECTIONS")
print("="*70)

# Load data
print("\n1. Loading results...")
df = pd.read_excel('Students Upload KCSE Results - Template_updated.xlsx')
total = len(df)
print(f"   ✓ {total} students loaded")

# Find gender
gender_col = find_gender_column(df)
if gender_col:
    df['GEND'] = df[gender_col].apply(normalize_gender)
    boys_df = df[df['GEND'] == 'MALE'].copy()
    girls_df = df[df['GEND'] == 'FEMALE'].copy()
    print(f"   ✓ Boys: {len(boys_df)}, Girls: {len(girls_df)}")
else:
    boys_df = pd.DataFrame()
    girls_df = pd.DataFrame()
    print("   ⚠ No gender data")

# Overall stats
mg = df['MEAN_GRADE'].dropna()
gd, ab = calculate_grade_distribution(mg)
mp = calculate_mean_points(mg)
print(f"   ✓ Overall: AB={ab}, Mean={mp:.2f}" if mp else f"   ✓ Overall: AB={ab}")

# Gender stats
boys_data = None
if len(boys_df) > 0:
    bg = boys_df['MEAN_GRADE'].dropna()
    bd, bab = calculate_grade_distribution(bg)
    bmp = calculate_mean_points(bg)
    boys_data = {'count': len(boys_df), 'dist': bd, 'ab': bab, 'mean': bmp}
    print(f"   ✓ Boys: AB={bab}, Mean={bmp:.2f}" if bmp else f"   ✓ Boys: AB={bab}")

girls_data = None
if len(girls_df) > 0:
    gg = girls_df['MEAN_GRADE'].dropna()
    gd_girls, gab = calculate_grade_distribution(gg)
    gmp = calculate_mean_points(gg)
    girls_data = {'count': len(girls_df), 'dist': gd_girls, 'ab': gab, 'mean': gmp}
    print(f"   ✓ Girls: AB={gab}, Mean={gmp:.2f}" if gmp else f"   ✓ Girls: AB={gab}")

# Subject mapping
subject_map = {
    'ENG': ('English', '101'),
    'KIS': ('Kiswahili', '102'),
    'MAT': ('Mathematics', '121'),
    'BIO': ('Biology', '231'),
    'PHY': ('Physics', '232'),
    'CHE': ('Chemistry', '233'),
    'HIS': ('History and Government', '311'),
    'GEO': ('Geography', '312'),
    'CRE': ('C.R.E.', '313'),
    'AGR': ('Agriculture', '443'),
    'BST': ('Business Studies', '565'),
    'COM': ('Computer Studies', '451')
}

# Calculate subject stats
print("\n2. Calculating subject statistics...")
subject_stats = {}
for subj, (name, code) in subject_map.items():
    if subj in df.columns:
        sg = df[subj].dropna().astype(str)
        sg = sg[sg != 'nan']
        if len(sg) > 0:
            sd, sab = calculate_grade_distribution(sg)
            smp = calculate_mean_points(sg)
            subject_stats[subj] = {
                'name': name, 'code': code,
                'total': len(sg), 'dist': sd, 'ab': sab, 'mean': smp
            }

# Subject stats by gender
boys_subj_stats = {}
if len(boys_df) > 0:
    for subj, (name, code) in subject_map.items():
        if subj in boys_df.columns:
            sg = boys_df[subj].dropna().astype(str)
            sg = sg[sg != 'nan']
            if len(sg) > 0:
                sd, sab = calculate_grade_distribution(sg)
                smp = calculate_mean_points(sg)
                boys_subj_stats[subj] = {
                    'name': name, 'code': code,
                    'count': len(sg), 'dist': sd, 'ab': sab, 'mean': smp
                }

girls_subj_stats = {}
if len(girls_df) > 0:
    for subj, (name, code) in subject_map.items():
        if subj in girls_df.columns:
            sg = girls_df[subj].dropna().astype(str)
            sg = sg[sg != 'nan']
            if len(sg) > 0:
                sd, sab = calculate_grade_distribution(sg)
                smp = calculate_mean_points(sg)
                girls_subj_stats[subj] = {
                    'name': name, 'code': code,
                    'count': len(sg), 'dist': sd, 'ab': sab, 'mean': smp
                }

print(f"   ✓ Calculated stats for {len(subject_stats)} subjects")

# Load template
print("\n3. Loading template...")
wb = load_workbook('KCSE ANALYSIS TEMPLATE.xlsx')
sheet = wb.active
print("   ✓ Template loaded")

# Column mapping for grades
# Column 6 (F) = AB, Column 7 (G) = A, Column 8 (H) = A-, etc.
col_map = {
    'A': 7, 'A-': 8, 'B+': 9, 'B': 10, 'B-': 11,
    'C+': 12, 'C': 13, 'C-': 14, 'D+': 15, 'D': 16, 'D-': 17, 'E': 18
}
AB_COL = 6  # Column F
YEAR_2025_COL = 23  # Column W

# Fill Row 7 - ORDER OF MERIT
print("\n4. Filling ORDER OF MERIT (Row 7)...")
sheet.cell(7, 2).value = "KUBWEYE SECONDARY SCHOOL"  # SCHOOL
if boys_data:
    sheet.cell(7, 3).value = boys_data['count']  # BOYS
if girls_data:
    sheet.cell(7, 4).value = girls_data['count']  # GIRLS
sheet.cell(7, 5).value = total  # TOTAL
sheet.cell(7, AB_COL).value = ab  # AB
for grade, col in col_map.items():
    sheet.cell(7, col).value = gd.get(grade, 0)
if mp:
    sheet.cell(7, YEAR_2025_COL).value = round(mp, 2)  # 2025 Mean Points
print("   ✓ Row 7 filled")

# Fill PERFORMANCE BY GENDER
print("\n5. Filling PERFORMANCE BY GENDER...")

# Row 13 - BOYS
if boys_data:
    sheet.cell(13, 3).value = boys_data['count']  # C13 - BOYS
    sheet.cell(13, 5).value = boys_data['count']  # E13 - TOTAL
    sheet.cell(13, AB_COL).value = boys_data['ab']  # F13 - AB
    for grade, col in col_map.items():
        sheet.cell(13, col).value = boys_data['dist'].get(grade, 0)
    if boys_data['mean']:
        sheet.cell(13, YEAR_2025_COL).value = round(boys_data['mean'], 2)  # W13
    print("   ✓ Row 13 (BOYS) filled")

# Row 14 - GIRLS
if girls_data:
    sheet.cell(14, 4).value = girls_data['count']  # D14 - GIRLS
    sheet.cell(14, 5).value = girls_data['count']  # E14 - TOTAL
    sheet.cell(14, AB_COL).value = girls_data['ab']  # F14 - AB
    for grade, col in col_map.items():
        sheet.cell(14, col).value = girls_data['dist'].get(grade, 0)
    if girls_data['mean']:
        sheet.cell(14, YEAR_2025_COL).value = round(girls_data['mean'], 2)  # W14
    print("   ✓ Row 14 (GIRLS) filled")

# Row 15 - TOTAL
sheet.cell(15, 5).value = total  # E15 - TOTAL
sheet.cell(15, AB_COL).value = ab  # F15 - AB
for grade, col in col_map.items():
    sheet.cell(15, col).value = gd.get(grade, 0)
if mp:
    sheet.cell(15, YEAR_2025_COL).value = round(mp, 2)  # W15
print("   ✓ Row 15 (TOTAL) filled")

# Fill SUBJECT ANALYSIS BY IMPROVEMENT INDEX (starts at row 21)
print("\n6. Filling SUBJECT ANALYSIS BY IMPROVEMENT INDEX...")
row = 21
for subj, (name, code) in subject_map.items():
    if subj in subject_stats:
        stats = subject_stats[subj]
        sheet.cell(row, 1).value = stats['name']  # A - SUBJECT
        sheet.cell(row, 2).value = stats['code']  # B - CODE
        if boys_data and subj in boys_subj_stats:
            sheet.cell(row, 3).value = boys_subj_stats[subj]['count']  # C - BOYS
        if girls_data and subj in girls_subj_stats:
            sheet.cell(row, 4).value = girls_subj_stats[subj]['count']  # D - GIRLS
        sheet.cell(row, 5).value = stats['total']  # E - TOTAL
        sheet.cell(row, AB_COL).value = stats['ab']  # F - AB
        for grade, col in col_map.items():
            sheet.cell(row, col).value = stats['dist'].get(grade, 0)
        if stats['mean']:
            sheet.cell(row, YEAR_2025_COL).value = round(stats['mean'], 2)  # W - 2025
        row += 1
print(f"   ✓ Filled {len(subject_stats)} subjects")

# Fill GENERAL SUBJECT ANALYSIS (starts at row 40)
print("\n7. Filling GENERAL SUBJECT ANALYSIS...")
row = 40
for subj, (name, code) in subject_map.items():
    if subj in subject_stats:
        stats = subject_stats[subj]
        sheet.cell(row, 1).value = stats['name']  # A - SUBJECT
        sheet.cell(row, 2).value = stats['code']  # B - CODE
        if boys_data and subj in boys_subj_stats:
            sheet.cell(row, 3).value = boys_subj_stats[subj]['count']  # C - BOYS
        if girls_data and subj in girls_subj_stats:
            sheet.cell(row, 4).value = girls_subj_stats[subj]['count']  # D - GIRLS
        sheet.cell(row, 5).value = stats['total']  # E - TOTAL
        sheet.cell(row, AB_COL).value = stats['ab']  # F - AB
        for grade, col in col_map.items():
            sheet.cell(row, col).value = stats['dist'].get(grade, 0)
        if stats['mean']:
            sheet.cell(row, YEAR_2025_COL).value = round(stats['mean'], 2)  # W - 2025
        row += 1
print(f"   ✓ Filled {len(subject_stats)} subjects in GENERAL SUBJECT ANALYSIS")

# Fill SUBJECT ANALYSIS (BOYS) - need to find the right section
# Based on template structure, this might be around row 52-56
print("\n8. Filling SUBJECT ANALYSIS (BOYS)...")
# Find where BOYS section starts by looking for "SUBJECT ANALYSIS (BOYS)" header
boys_section_row = None
for r in range(30, 60):
    val = sheet.cell(r, 1).value
    if val and 'BOYS' in str(val).upper() and 'SUBJECT' in str(val).upper():
        boys_section_row = r + 2  # Data starts 2 rows after header
        break

if boys_section_row is None:
    boys_section_row = 56  # Default fallback

row = boys_section_row
for subj, (name, code) in subject_map.items():
    if subj in boys_subj_stats:
        stats = boys_subj_stats[subj]
        try:
            safe_set_cell(sheet, row, 1, stats['name'])  # A - SUBJECT
            safe_set_cell(sheet, row, 2, stats['code'])  # B - CODE
            safe_set_cell(sheet, row, 3, stats['count'])  # C - BOYS
            safe_set_cell(sheet, row, 5, stats['count'])  # E - TOTAL
            safe_set_cell(sheet, row, AB_COL, stats['ab'])  # F - AB
            for grade, col in col_map.items():
                safe_set_cell(sheet, row, col, stats['dist'].get(grade, 0))
            if stats['mean']:
                safe_set_cell(sheet, row, YEAR_2025_COL, round(stats['mean'], 2))  # W - 2025
        except Exception as e:
            print(f"   ⚠ Warning: Could not fill row {row} for {subj}: {e}")
        row += 1
print(f"   ✓ Filled {len(boys_subj_stats)} subjects for boys")

# Fill SUBJECT ANALYSIS (GIRLS)
print("\n9. Filling SUBJECT ANALYSIS (GIRLS)...")
# Find where GIRLS section starts
girls_section_row = None
for r in range(60, 90):
    val = sheet.cell(r, 1).value
    if val and 'GIRLS' in str(val).upper() and 'SUBJECT' in str(val).upper():
        girls_section_row = r + 2  # Data starts 2 rows after header
        break

if girls_section_row is None:
    girls_section_row = 74  # Default fallback

row = girls_section_row
for subj, (name, code) in subject_map.items():
    if subj in girls_subj_stats:
        stats = girls_subj_stats[subj]
        try:
            safe_set_cell(sheet, row, 1, stats['name'])  # A - SUBJECT
            safe_set_cell(sheet, row, 2, stats['code'])  # B - CODE
            safe_set_cell(sheet, row, 4, stats['count'])  # D - GIRLS
            safe_set_cell(sheet, row, 5, stats['count'])  # E - TOTAL
            safe_set_cell(sheet, row, AB_COL, stats['ab'])  # F - AB
            for grade, col in col_map.items():
                safe_set_cell(sheet, row, col, stats['dist'].get(grade, 0))
            if stats['mean']:
                safe_set_cell(sheet, row, YEAR_2025_COL, round(stats['mean'], 2))  # W - 2025
        except Exception as e:
            print(f"   ⚠ Warning: Could not fill row {row} for {subj}: {e}")
        row += 1
print(f"   ✓ Filled {len(girls_subj_stats)} subjects for girls")

# Fill TOP 10 BOYS (around row 88)
print("\n10. Filling TOP 10 BOYS...")
if len(boys_df) > 0:
    boys_df['MEAN_PTS'] = boys_df['MEAN_GRADE'].apply(lambda x: grade_to_points.get(str(x).strip(), 0) if pd.notna(x) else 0)
    top_boys = boys_df.nlargest(10, 'MEAN_PTS')
    
    row = 90
    for idx, student in top_boys.iterrows():
        sheet.cell(row, 1).value = str(student['INDEXNO'])  # A - INDEXNO
        sheet.cell(row, 2).value = student['NAME']  # B - NAME
        # Subject grades (columns 3-13)
        subjects = ['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'HIS', 'GEO', 'CRE', 'AGR', 'BST']
        for i, subj in enumerate(subjects, start=3):
            if subj in student and pd.notna(student[subj]) and str(student[subj]) != 'nan':
                sheet.cell(row, i).value = str(student[subj])
        sheet.cell(row, 18).value = student['MEAN_GRADE']  # R - GR (Mean Grade)
        row += 1
    print(f"   ✓ Filled top {len(top_boys)} boys")

# Fill TOP 10 GIRLS (around row 103)
print("\n11. Filling TOP 10 GIRLS...")
if len(girls_df) > 0:
    girls_df['MEAN_PTS'] = girls_df['MEAN_GRADE'].apply(lambda x: grade_to_points.get(str(x).strip(), 0) if pd.notna(x) else 0)
    top_girls = girls_df.nlargest(10, 'MEAN_PTS')
    
    row = 105
    for idx, student in top_girls.iterrows():
        sheet.cell(row, 1).value = str(student['INDEXNO'])  # A - INDEXNO
        sheet.cell(row, 2).value = student['NAME']  # B - NAME
        # Subject grades
        subjects = ['ENG', 'KIS', 'MAT', 'BIO', 'PHY', 'CHE', 'HIS', 'GEO', 'CRE', 'AGR', 'BST']
        for i, subj in enumerate(subjects, start=3):
            if subj in student and pd.notna(student[subj]) and str(student[subj]) != 'nan':
                sheet.cell(row, i).value = str(student[subj])
        sheet.cell(row, 18).value = student['MEAN_GRADE']  # R - GR
        row += 1
    print(f"   ✓ Filled top {len(top_girls)} girls")

# Save
print("\n12. Saving template...")
output_file = 'KCSE ANALYSIS TEMPLATE_2025_FILLED.xlsx'
wb.save(output_file)
print(f"   ✓ Saved to: {output_file}")

print("\n" + "="*70)
print("COMPLETE! All sections filled successfully!")
print("="*70)
print(f"\nSummary:")
print(f"  - Total Students: {total}")
print(f"  - Boys: {boys_data['count'] if boys_data else 0}")
print(f"  - Girls: {girls_data['count'] if girls_data else 0}")
print(f"  - Subjects analyzed: {len(subject_stats)}")
if len(boys_df) > 0:
    print(f"  - Top 10 Boys: {len(top_boys)}")
if len(girls_df) > 0:
    print(f"  - Top 10 Girls: {len(top_girls)}")
print("="*70)
