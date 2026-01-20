#!/usr/bin/env python3
"""Fill gender sections in KCSE Analysis Template - Standalone version."""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

# Grade to points
grade_to_points = {'A': 12, 'A-': 11, 'B+': 10, 'B': 9, 'B-': 8, 'C+': 7, 'C': 6, 'C-': 5, 'D+': 4, 'D': 3, 'D-': 2, 'E': 1}

def calc_grade_dist(grades):
    g = grades.dropna().astype(str)
    dist = {}
    for grade in ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']:
        dist[grade] = len(g[g == grade])
    ab = dist.get('A', 0) + dist.get('A-', 0) + dist.get('B+', 0) + dist.get('B', 0) + dist.get('B-', 0)
    return dist, ab

def calc_mean_pts(grades):
    g = grades.dropna().astype(str)
    pts = [grade_to_points[x] for x in g if x in grade_to_points]
    return np.mean(pts) if pts else None

# Load data
df = pd.read_excel('Students Upload KCSE Results - Template_updated.xlsx')

# Find GENDER column
gender_col = None
for col in df.columns:
    if 'GENDER' in str(col).upper():
        gender_col = col
        break

if not gender_col:
    with open('gender_fill_status.txt', 'w') as f:
        f.write("ERROR: GENDER column not found!\n")
        f.write(f"Columns: {list(df.columns)}\n")
    exit(1)

# Normalize gender
def norm_gender(v):
    if pd.isna(v):
        return None
    v = str(v).strip().upper()
    return 'MALE' if v in ['M', 'MALE', 'BOY', 'BOYS'] else 'FEMALE' if v in ['F', 'FEMALE', 'GIRL', 'GIRLS'] else None

df['GEND_NORM'] = df[gender_col].apply(norm_gender)

# Calculate stats
boys_df = df[df['GEND_NORM'] == 'MALE']
girls_df = df[df['GEND_NORM'] == 'FEMALE']

boys_data = None
if len(boys_df) > 0:
    bg = boys_df['MEAN_GRADE'].dropna()
    bd, bab = calc_grade_dist(bg)
    bmp = calc_mean_pts(bg)
    boys_data = {'count': len(boys_df), 'dist': bd, 'ab': bab, 'mean': bmp}

girls_data = None
if len(girls_df) > 0:
    gg = girls_df['MEAN_GRADE'].dropna()
    gd, gab = calc_grade_dist(gg)
    gmp = calc_mean_pts(gg)
    girls_data = {'count': len(girls_df), 'dist': gd, 'ab': gab, 'mean': gmp}

# Load template
wb = load_workbook('KCSE ANALYSIS TEMPLATE.xlsx')
sheet = wb.active

col_map = {'A': 7, 'A-': 8, 'B+': 9, 'B': 10, 'B-': 11, 'C+': 12, 'C': 13, 'C-': 14, 'D+': 15, 'D': 16, 'D-': 17, 'E': 18}

# Fill Row 7
if boys_data:
    sheet.cell(7, 3).value = boys_data['count']
if girls_data:
    sheet.cell(7, 4).value = girls_data['count']

# Fill Row 13 (BOYS)
if boys_data:
    sheet.cell(13, 3).value = boys_data['count']  # C13
    sheet.cell(13, 5).value = boys_data['count']  # E13
    sheet.cell(13, 6).value = boys_data['ab']     # F13
    for grade, col in col_map.items():
        sheet.cell(13, col).value = boys_data['dist'].get(grade, 0)
    if boys_data['mean']:
        sheet.cell(13, 23).value = round(boys_data['mean'], 2)

# Fill Row 14 (GIRLS)
if girls_data:
    sheet.cell(14, 4).value = girls_data['count']  # D14
    sheet.cell(14, 5).value = girls_data['count']  # E14
    sheet.cell(14, 6).value = girls_data['ab']     # F14
    for grade, col in col_map.items():
        sheet.cell(14, col).value = girls_data['dist'].get(grade, 0)
    if girls_data['mean']:
        sheet.cell(14, 23).value = round(girls_data['mean'], 2)

# Save
wb.save('KCSE ANALYSIS TEMPLATE_2025_FILLED.xlsx')

# Write status
with open('gender_fill_status.txt', 'w') as f:
    f.write("Gender sections filled successfully!\n")
    f.write(f"Boys: {boys_data['count'] if boys_data else 0}\n")
    f.write(f"Girls: {girls_data['count'] if girls_data else 0}\n")
    f.write(f"Row 13 C13: {sheet.cell(13, 3).value}\n")
    f.write(f"Row 14 D14: {sheet.cell(14, 4).value}\n")

print("Gender sections filled! Check gender_fill_status.txt for details.")
